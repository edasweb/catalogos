
# generate bue one

import openpyxl
from datetime import date
from openpyxl.styles import Font

from pprint import pprint


brands = ["VW", "Acura", "Volkswagen", "BMW", "Audi", "MINI", "Mitsubish", "Honda","Hyundai","Mercedes","Chevrolet","Cadillac","Buick","Ford",
          "Toyota","Subaru","SEAT","All Models", "Mazda", "Porsche", "Fiat", "Dodge", "jeep", "Suzuki"]

wb = openpyxl.load_workbook("base.xlsx") 



          
manufacturer = "PRACWORKS"

sheet_name = "base"

# print(wb.sheetnames) 
sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"catalogue - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


color_sheet_name = "colors"
color_sheet = wb[color_sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
colors = []
for row in range(1, color_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    colors.append(color_sheet[f"a{row}"].value)



rail_sheet_name = "rails"
rail_sheet = wb[rail_sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
rails = []
for row in range(1, rail_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    rails.append(rail_sheet[f"a{row}"].value)

finish_sheet_name = "finishs"
finish_sheet = wb[finish_sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
finishs = []
for row in range(1, finish_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    finishs.append(finish_sheet[f"a{row}"].value)




temp_row = dict()


category_name = None
for row in range(2, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    if category_name != sheet_1[f"e{row}"].value.upper():
        category_name = sheet_1[f"e{row}"].value.upper()
        sheet_2.append([category_name.upper()])
        sheet_2.append(["Car Application", "Part Number", "Description","Retail Price Ex VAT €", "Application" ])


    reference = sheet_1[f"b{row}"].value if sheet_1[f"b{row}"].value != None else ""
    


    
    name = sheet_1[f"a{row}"].value
    part_number = sheet_1[f"b{row}"].value
    description = sheet_1[f"c{row}"].value
    price = sheet_1[f"d{row}"].value
    application = sheet_1[f"a{row}"].value + " " + description if description != None else ""
    sheet_2.append([name, part_number, description, price,application ])             
    
  
    
wb.save(manufacturer.upper() + "_asd_catalogue_copy".upper() + ".xlsx")


