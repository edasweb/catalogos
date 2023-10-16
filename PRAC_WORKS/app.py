
# generate bue one

import openpyxl
from datetime import date
from openpyxl.styles import Font

from pprint import pprint


brands = ["VW", "Acura", "Volkswagen", "BMW", "Audi", "MINI", "Mitsubish", "Honda","Hyundai","Mercedes","Chevrolet","Cadillac","Buick","Ford",
          "Toyota","Subaru","SEAT","All Models", "Mazda", "Porsche", "Fiat", "Dodge", "jeep", "Suzuki"]

wb = openpyxl.load_workbook("base.xlsx") 


def short_description(reference):
    if "HK-N-20" in reference:
        return "HK-N-20"
    if "HK-L-20" in reference:
        return "HK-L-20"
    if "HK-N-CTR" in reference:
        return "HK-N-CTR"
    if "HK-R-00" in reference:
        return "HK-R-00"
    if "HB-N-20" in reference:
        return "HB-N-20"
    if "HF-N-00" in reference:
        return "HF-N-00"
    if "HX-N-00" in reference:
        return "HX-N-00"
    if "HK-V-01" in reference:
        return "HK-V-01"
    if "UFP-6-3.25" in reference:
        return "UFP-6-3.25"
    return "No REFERENCE"
    




          
manufacturer = "PRACWORKS"

sheet_name = "base"

# print(wb.sheetnames) 
sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"catalogue - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet

prices = dict()
color_sheet_name = "colors"
color_sheet = wb[color_sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
colors = []
for row in range(1, 3): # To iterate over all the row and column of the sheet and get each value.
    colors.append(color_sheet[f"a{row}"].value)
    prices[color_sheet[f"a{row}"].value] = color_sheet[f"b{row}"].value


rail_sheet_name = "rails"
rail_sheet = wb[rail_sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
rails = []

for row in range(1, 3): # To iterate over all the row and column of the sheet and get each value.
    rails.append(rail_sheet[f"a{row}"].value)
    prices[rail_sheet[f"a{row}"].value] = rail_sheet[f"b{row}"].value
finish_sheet_name = "finishs"
finish_sheet = wb[finish_sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
finishs = []
for row in range(1, 3): # To iterate over all the row and column of the sheet and get each value.
    finishs.append(finish_sheet[f"a{row}"].value)
    prices[finish_sheet[f"a{row}"].value] = finish_sheet[f"b{row}"].value




temp_row = dict()


category_name = None
for row in range(1, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    

    reference = sheet_1[f"b{row}"].value if sheet_1[f"b{row}"].value != None else ""
    


    reference = reference.replace("(", "").replace(")", "")
    
    if "color" in reference:
        
        for color in colors:
            
            name = sheet_1[f"a{row}"].value
            
            part_number = sheet_1[f"b{row}"].value.replace("(", "").replace(")", "")
            pprint(part_number)
            description = sheet_1[f"c{row}"].value
            price = sheet_1[f"d{row}"].value
            if "finish" in reference:
                for finish in finishs:
                    if "rail" in reference:
                        for rail in rails:
                            print(row)
                            name = name + " " + color[0].upper() + color[1:].lower() + " " + finish[0].upper() + finish[1:].lower() + " " + (rail[0].upper() + rail[1:].lower() if rail != "none" else "No Rail")
                            part_number = part_number.replace("color", color).replace("finish", finish).replace("rail", rail).upper().replace("(", "").replace(")", "")
                            price = price + prices[color] + prices[finish] + prices[rail]
                            sheet_2.append([name.replace(" None", " No Rail"), part_number.replace("BLACK", "BL").replace("RED", "RD").replace("NONE", "NO").replace("POLISHED", "PO").replace("RAIL", "RA").replace("MATTE", "MA"), description, price, short_description(reference)])
                            name = sheet_1[f"a{row}"].value
                            part_number = sheet_1[f"b{row}"].value
                            price = sheet_1[f"d{row}"].value
                    else:
                       name = name + " " + color[0].upper() + color[1:].lower() + " " + finish[0].upper() + finish[1:].lower()
                       part_number = part_number.replace("color", color).replace("finish", finish).upper().replace("(", "").replace(")", "")
                       price = price + prices[color] + prices[finish] 

                       sheet_2.append([name.replace(" None", " No Rail"), part_number.replace("BLACK", "BL").replace("RED", "RD").replace("NONE", "NO").replace("POLISHED", "PO").replace("RAIL", "RA").replace("MATTE", "MA"), description, price, short_description(reference)]) 
                       name = sheet_1[f"a{row}"].value
                       part_number = sheet_1[f"b{row}"].value
                       price = sheet_1[f"d{row}"].value
            else:
               name = name + " " + color[0].upper() + color[1:].lower()
               part_number = part_number.replace("color", color).upper().replace("(", "").replace(")", "")
               price = price + prices[color]  
               sheet_2.append([name.replace(" None", " No Rail"), part_number.replace("BLACK", "BL").replace("RED", "RD").replace("NONE", "NO").replace("POLISHED", "PO").replace("RAIL", "RA").replace("MATTE", "MA"), description, price, short_description(reference)]) 
               name = sheet_1[f"a{row}"].value
               part_number = sheet_1[f"b{row}"].value 
               price = sheet_1[f"d{row}"].value
    else:
        
        name = sheet_1[f"a{row}"].value
        part_number = sheet_1[f"b{row}"].value.upper().replace("(", "").replace(")", "") if sheet_1[f"b{row}"].value != None else ""
        description = sheet_1[f"c{row}"].value
        price = sheet_1[f"d{row}"].value
        sheet_2.append([name.replace(" None", " No Rail") if name != None else name, part_number.replace("BLACK", "BL").replace("RED", "RD").replace("NONE", "NO").replace("POLISHED", "PO").replace("RAIL", "RA").replace("MATTE", "MA"), description, price, short_description(reference)])             
    
  
    
wb.save(manufacturer.upper() + "_asd_catalogue5".upper() + ".xlsx")


