

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s


base_sheet_name = "prices_base"
wb = openpyxl.load_workbook("base.xlsx") 

base_sheet = wb[base_sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
import_sheet_name = "import"
import_sheet = wb.create_sheet(title=f"{import_sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


manufacturer = "PracWorks"
import_sheet.append(["Manufacturer","Supplier","Price","Purchase","%","VAT","Reference","Name","Name","EAN13","Category","Meta title","Tags","Keywords","Rewrite","shor_description",
                      
])





for row in range(2, 66): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()

    if base_sheet[f"b{row}"].value != None:
    
      reference = base_sheet[f"b{row}"].value
      
      temp_row = [
            manufacturer,                                # "manufacturer"
            manufacturer,                                # "Supplier"
            base_sheet[f"d{row}"].value,                   # "price"
            base_sheet[f"d{row}"].value * Decimal(0.65),                 # "purchase"
            "25",                                       # " % "
            "7",                                        # "VAT"
            reference,                   # "reference"
            manufacturer + " " + base_sheet[f"a{row}"].value+ " ", # "name"
            manufacturer + " " + base_sheet[f"a{row}"].value+ " " + base_sheet[f"c{row}"].value, # "name"
            "",                   # "EAN13"
            "Home",                                     # "category"
            manufacturer + " " + base_sheet[f"a{row}"].value+ " ",                   # "Meta title"
            manufacturer + ","+reference  + ","+ str(base_sheet[f"a{row}"].value + "," + base_sheet[f"c{row}"].value).replace(" ", ",").replace("-", ",").replace("/", ","),                   # "Tags"
            manufacturer + ","+ reference  +","+ str(base_sheet[f"a{row}"].value + "," +base_sheet[f"c{row}"].value).replace(" ", ",").replace("-", ",").replace("/", ","),                   # "Keywords"
            slugify(manufacturer + "-"+ reference),                   # "rewrite"
            base_sheet[f"e{row}"].value,
            
            
        ]
      import_sheet.append(temp_row)
    
wb.save(manufacturer.upper() + "_ALL_STARS_IMPORT_CATALOGUE_FILTERED2.xlsx")






