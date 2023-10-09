

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s




wb = openpyxl.load_workbook("alfa_with_variations3.xlsx") 

# print(wb.sheetnames) 



chassi_sheet = wb.create_sheet(title=f"chassis- {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet

sheet_name_brand_together = "brand-together"




sheet_brand_together = wb[sheet_name_brand_together] # To accsses the a sheet in the workbook. And create a sheet object.

chassi_name = None
chassi_sheet.append(["Refrence", "Brand","Model","Chassis","Type"])

for row in range(1, sheet_brand_together.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    if chassi_name !=str(sheet_brand_together[f"d{row}"].value) and "..." in str(sheet_brand_together[f"d{row}"].value):
        chassi_name = str(sheet_brand_together[f"d{row}"].value)

        chassi_sheet.append([sheet_brand_together[f"a{row}"].value, 
                             sheet_brand_together[f"b{row}"].value, 
                             sheet_brand_together[f"c{row}"].value,
                             sheet_brand_together[f"d{row}"].value,
                             sheet_brand_together[f"e{row}"].value,
                             ])
    

wb.save("alfa_chassis.xlsx")






