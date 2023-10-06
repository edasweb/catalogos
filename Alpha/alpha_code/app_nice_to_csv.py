

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s




wb = openpyxl.load_workbook("alfa_with_variations3.xlsx") 
sheet_price_name = "prices"
to_csv_sheet = wb.create_sheet(title=f"nice- {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet

# print(wb.sheetnames) 

sheet_prices = wb[sheet_price_name] # To accsses the a sheet in the workbook. And create a sheet object.

prices = {}
references = list()
for row in range(2, sheet_prices.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    reference = sheet_prices[f"h{row}"].value.upper().strip().replace(" ", "")
    

    if prices.get(reference) == None and reference != None:
        references.append(reference)
        prices[reference] = {
            "price":sheet_prices[f"c{row}"].value, 
                             "wholesale_price":sheet_prices[f"d{row}"].value,
                             "attr_price":sheet_prices[f"e{row}"].value,
                             "attr_wholesale_price":sheet_prices[f"f{row}"].value,
                             "name":sheet_prices[f"g{row}"].value,
                             }




to_csv_sheet.append(["Manufacturer","Supplier","Price EUR","Purchase","Discount %","VAT","Reference","Name","Category","Meta title","Tags","Keywords","Rewrite"])


for key in prices:
    if key == "AC-MK7-DI.A":
        print(prices[key])
    temp_row = list()
    detail = prices[key]
    temp_row = [
        "Alpha Competition",                                # "manufacturer"
        "Alpha Competition",                                # "Supplier"
        detail["price"],                   # "price"
        detail["wholesale_price"], # "purchase"
        "--descont--",                                       # " % "
        "7",                                        # "VAT"
        key,                   # "reference"
        detail["name"],  # "name"
        "Home",                   # "category"
        detail["name"],    # "Meta title"
        "Alpha Competition," + key  +","+detail["name"].replace(" ", ","),                   # "Tags"
        "Alpha Competition," + key  +","+detail["name"].replace(" ", ","),             # "Keywords"
        slugify("Alpha-Competition-" + key),                   # "rewrite"
       
    ]
    to_csv_sheet.append(temp_row)
        

    
wb.save("alpha_competition_imports.xlsx")






