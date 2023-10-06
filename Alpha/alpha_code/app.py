
# generate bue one

import openpyxl
from datetime import date


sheet_name = "all"
wb = openpyxl.load_workbook("separated_new_update.xlsx") 
# print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


def getString(value):
    if value:
        return str(value)
    
    return ""

brand_name = None
group = None
for row in range(1, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    if brand_name !=str(sheet_1[f"j{row}"].value).upper():
        brand_name = str(sheet_1[f"j{row}"].value).upper()
        sheet_2.append([brand_name])
        sheet_2.append(["Description","Reference", "Retail Price Ex VAT €","Length","Width","Height","Weight"
])
    temp_row = [
        sheet_1[f"c{row}"].value, 
        sheet_1[f"a{row}"].value, 
        sheet_1[f"b{row}"].value, 
        sheet_1[f"d{row}"].value, 
        sheet_1[f"e{row}"].value, 
        sheet_1[f"f{row}"].value, 
        sheet_1[f"g{row}"].value, 
       
    ]
    sheet_2.append(temp_row)
    


wb.save("p3_catalogue_new.xlsx")


