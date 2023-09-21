
# generate bue one

import openpyxl
from datetime import date


sheet_name = "Worksheet"
wb = openpyxl.load_workbook("Work_File.xlsx") 
print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


def getString(value):
    if value:
        return str(value)
    
    return ""

brand_name = None
group = None
for row in range(1, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    if brand_name !=str(sheet_1[f"b{row}"].value).upper():
        brand_name = str(sheet_1[f"b{row}"].value).upper()
        sheet_2.append([brand_name , "", "", "", "", "", "", "", ""])
        sheet_2.append(["Application", "Description","Reference", "Retail Price Ex VAT €","Fits To","Pipe Diameter","Tailpipes", "Valve","Note"])
    print(row)

    # temp_row = [
    #     sheet_1[f"b{row}"].value + " " + sheet_1[f"c{row}"].value+ " " + (str(sheet_1[f"f{row}"].value) + " - " + str(sheet_1[f"g{row}"].value)) if sheet_1[f"f{row}"].value and sheet_1[f"f{row}"].value else "", 
    #     sheet_1[f"d{row}"].value, 
    #     sheet_1[f"b{row}"].value + " " + sheet_1[f"c{row}"].value+ " " + getString(sheet_1[f"l{row}"].value) + " ", 
    #     sheet_1[f"b{row}"].value + " " + sheet_1[f"c{row}"].value+ " " + getString(sheet_1[f"l{row}"].value) + " "+ (" PD(" if sheet_1[f"k{row}"].value != None else "") + getString(sheet_1[f"k{row}"].value)+ (")" if sheet_1[f"k{row}"].value != None else "")   + " "+ (" TD(" if sheet_1[f"m{row}"].value != None else "") + getString(sheet_1[f"m{row}"].value) + (")" if sheet_1[f"m{row}"].value != None else "") , 
    #     sheet_1[f"j{row}"].value, 
    #     sheet_1[f"k{row}"].value, 
    #     sheet_1[f"m{row}"].value, 
    #     sheet_1[f"h{row}"].value, 
    # ]
    temp_row = [
        sheet_1[f"c{row}"].value + " ", 
        sheet_1[f"d{row}"].value, 
        sheet_1[f"e{row}"].value, 
        sheet_1[f"h{row}"].value, 
        sheet_1[f"j{row}"].value, 
        sheet_1[f"k{row}"].value, 
        sheet_1[f"m{row}"].value, 
        sheet_1[f"o{row}"].value, 
        sheet_1[f"p{row}"].value, 
    ]
    sheet_2.append(temp_row)
    


wb.save("scorpion_catalogue_new.xlsx")


