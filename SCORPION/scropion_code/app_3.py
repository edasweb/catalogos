

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s


sheet_name = "Worksheet"
wb = openpyxl.load_workbook("Work_File.xlsx") 
print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


brand_name = None


sheet_2.append(["Manufacturer","Price","Reference","Fits To",	"Pipe Diameter","Tailpipe",	"Tailpipe Diameter",		"Valve",	"Notes"
])

for row in range(2, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    temp_row = [
        "Scorpion",                                # "manufacturer"
        sheet_1[f"h{row}"].value,                   # "price"
        sheet_1[f"e{row}"].value,                   # "reference"
        sheet_1[f"j{row}"].value,                                         # "Fits To"
        sheet_1[f"k{row}"].value,                                         # "Pipe Diameter"
        sheet_1[f"l{row}"].value,                                         # "Tailpipe"
        sheet_1[f"m{row}"].value,                                         # "Tailpipe Diameter"
        sheet_1[f"o{row}"].value,                                         # "Valve"
        sheet_1[f"p{row}"].value,                                         # "Notes"
    ]
    sheet_2.append(temp_row)
    
wb.save("templace_csv_clientes_new.xlsx")






