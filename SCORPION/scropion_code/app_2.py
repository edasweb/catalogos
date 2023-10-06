

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s

def getString(value):
    if value:
        return " "+str(value)
    
    return ""


sheet_name = "Worksheet"
wb = openpyxl.load_workbook("Work_File.xlsx") 
# print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


brand_name = None
sheet_2.append(["Manufacturer","Supplier","Price","Purchase","%","VAT","Reference","Name","Category","Meta title","Tags","Keywords","Rewrite"])

for row in range(2, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    temp_row = [
        "Scorpion",                                # "manufacturer"
        sheet_1[f"a{row}"].value,                                # "Supplier"
        sheet_1[f"h{row}"].value,                   # "price"
        Decimal(sheet_1[f"h{row}"].value) * Decimal(0.50), # "purchase"
        "35",                                       # " % "
        "7",                                        # "VAT"
        sheet_1[f"e{row}"].value,                   # "reference"
        sheet_1[f"a{row}"].value + " " + sheet_1[f"c{row}"].value + " " + sheet_1[f"d{row}"].value,  # "name"
        "Home",                   # "category"
        sheet_1[f"a{row}"].value + " " + sheet_1[f"c{row}"].value+ " " + sheet_1[f"d{row}"].value,    # "Meta title"
        "Scorpion," + sheet_1[f"e{row}"].value + "," +sheet_1[f"b{row}"].value + "," + sheet_1[f"c{row}"].value + "," + getString(sheet_1[f"j{row}"].value) + "," + getString(sheet_1[f"d{row}"].value)+ ","+ getString(sheet_1[f"l{row}"].value) + ","+ getString(sheet_1[f"k{row}"].value)+  "," + getString(sheet_1[f"m{row}"].value),                   # "Tags"
        "Scorpion," + sheet_1[f"e{row}"].value + "," +sheet_1[f"b{row}"].value.replace(" ", ",") + "," + sheet_1[f"c{row}"].value.replace(" ", ",") + "," + getString(sheet_1[f"j{row}"].value).replace(" ", ",") + "," + sheet_1[f"d{row}"].value.replace(" ", ",")+ ","+ getString(sheet_1[f"l{row}"].value).replace(" ", ",") + ","+ getString(sheet_1[f"k{row}"].value).replace(" ", ",")+  "," + getString(sheet_1[f"m{row}"].value).replace(" ", ",")  ,             # "Keywords"
        slugify("Scorpion-" + sheet_1[f"e{row}"].value),                   # "rewrite"
       
    ]
    sheet_2.append(temp_row)
    
wb.save("scorpion_all_stars_import_catalogue_new29.xlsx")






