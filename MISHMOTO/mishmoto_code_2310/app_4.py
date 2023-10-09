

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s


sheet_name = "Prices - 31 August 2023 - Sep-2"
wb = openpyxl.load_workbook("templace_csv_clientes_new.xlsx") 
print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


brand_name = None




for row in range(1, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    temp_row = [
        
        sheet_1[f'a{row}'].value + ','+str(sheet_1[f'b{row}'].value) + ','+sheet_1[f'c{row}'].value+ ','+sheet_1[f'd{row}'].value + ','+ (sheet_1[f'e{row}'].value if sheet_1[f'e{row}'].value != None else '') + ',"' + sheet_1[f"f{row}"].value +'"'                
        
        
    ]
    sheet_2.append(temp_row)
    
wb.save("templace_csv_clientes_new_created.xlsx")






