
# generate bue one

import openpyxl
from datetime import date
from openpyxl.styles import Font

from pprint import pprint


brands = ["VW", "Acura", "Volkswagen", "BMW", "Audi", "MINI", "Mitsubish", "Honda","Hyundai","Mercedes","Chevrolet","Cadillac","Buick","Ford",
          "Toyota","Subaru","SEAT","All Models", "Mazda", "Porsche", "Fiat", "Dodge", "jeep", "Suzuki"]





          

sheet_name = "base"
wb = openpyxl.load_workbook("base.xlsx") 
# print(wb.sheetnames) 
import operator
sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"nice - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet

manufacturer = "Mishimoto"
category_name = None
temp_row = dict()

def sort_key(row):
    return row[2]




for row in range(2, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    
    if category_name != sheet_1[f"c{row}"].value:
        category_name = sheet_1[f"c{row}"].value
        
    

    if temp_row.get(category_name.strip()) == None:
        temp_row[category_name.strip()] = list()
    
    temp_row[category_name].append( [
            sheet_1[f"a{row}"].value, 
            sheet_1[f"b{row}"].value, 
            sheet_1[f"e{row}"].value, 
            sheet_1[f"f{row}"].value, 
            sheet_1[f"g{row}"].value, 
            sheet_1[f"h{row}"].value, 
            sheet_1[f"j{row}"].value, 
        ])   
category_name = ""   
for category in temp_row:
    # pprint(category)
    if category_name != category:
        category_name = category
        sheet_2.append([category_name.upper()])
        sheet_2.append(["Part Number", "Description", "Weight", "Height", "Width", "Lenght","Retail Price Ex VAT €" ])

    if temp_row[category] != None:
        # print(temp_row[category])
        for row in sorted(temp_row[category], key=lambda x: x[1]) :
            # pprint(row)
            sheet_2.append(row)
    


wb.save(manufacturer.upper() + "_asd_catalogue.xlsx".upper())


