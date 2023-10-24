

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s

def getString(value):
    if value:
        return " "+str(value)
    
    return ""


sheet_name = "website"
wb = openpyxl.load_workbook("base.xlsx") 
# print(wb.sheetnames) 

web_site_sheet = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.



removed_sheet_name = "removed"
removed_sheet = wb[removed_sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
removed_references = list()

for row in range(1, removed_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    
      removed_references.append(removed_sheet[f"a{row}"].value)






print(removed_references)

query_sheet = wb.create_sheet(title=f"queries - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


brand_name = None
query_sheet.append(["ID","Reference", "Discount","Query"])

for row in range(1, web_site_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    if web_site_sheet[f"c{row}"].value != None and not web_site_sheet[f"c{row}"].value in removed_references:
        query =  f""" INSERT INTO `psnz_specific_price`
        ( `id_specific_price_rule`, `id_cart`, `id_product`, `id_shop`, `id_shop_group`, `id_currency`, `id_country`, `id_group`, `id_customer`, `id_product_attribute`, `price`, `from_quantity`, `reduction`, `reduction_tax`, `reduction_type`, `from`, `to`) VALUES 
        (0,0,{web_site_sheet[f"a{row}"].value},0,0,0,0,0,0,0,-1,1,0.3,1,"percentage","0000-00-00 00:00:00","0000-00-00 00:00:00") 
        """
        temp_row = [
                                        
            web_site_sheet[f"a{row}"].value,                                
            web_site_sheet[f"c{row}"].value,                   
            30,                   
            query,                    
        ]
        query_sheet.append(temp_row)
wb.save("QUERYS.xlsx")

# print(all_query_toguether)






