

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s
from pprint import pprint

wb = openpyxl.load_workbook("base.xlsx") 
manufacturer = "Mishimoto"
# print(wb.sheetnames) 



sheet_name = "website_now"
website_sheet = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
website_references = list()

for row in range(1, website_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    if website_sheet[f"c{row}"].value != None: 
      website_references.append(website_sheet[f"c{row}"].value.strip())


sheet_name = "prices"
price_sheet = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
price_references = list()

for row in range(1, price_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    price_references.append(price_sheet[f"g{row}"].value.strip())


sheet_name = "old_but_on_stock"
price_sheet_stock = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

for row in range(1, price_sheet_stock.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    price_references.append(price_sheet_stock[f"a{row}"].value.strip())



new_sheet = wb.create_sheet(title=f"figure_out - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet

for row in range(1, website_sheet.max_row + 1):
    if website_sheet[f"c{row}"].value != None:
        reference = website_sheet[f"c{row}"].value.strip()
        if not reference in price_references:
            new_sheet.append([reference])
        





        

  
    
wb.save(manufacturer.upper() + "_figure_out.xlsx".upper())






