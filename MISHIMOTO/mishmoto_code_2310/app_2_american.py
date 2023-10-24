

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s




base_sheet_name = "base"
wb = openpyxl.load_workbook("base.xlsx") 

base_sheet = wb[base_sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
import_sheet_name = "import"
import_sheet = wb.create_sheet(title=f"{import_sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


manufacturer = "Mishimoto"
import_sheet.append(["Manufacturer","Supplier","Price","Purchase","%","VAT","Reference","Name","EAN13","Category","Meta title","Tags","Keywords","Rewrite",
                      "Weight","Height","Width","Depth"
])

for row in range(2, base_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    if "Ford Mustang".upper() in base_sheet[f"b{row}"].value.upper()  or "Chevrolet".upper() in base_sheet[f"b{row}"].value.upper():
      temp_row = [
          manufacturer,                                # "manufacturer"
          manufacturer,                                # "Supplier"
          base_sheet[f"j{row}"].value,                 # "price"
          base_sheet[f"k{row}"].value,                 # "purchase"
          "30",                                        # " % "
          "7",                                          # "VAT"
          base_sheet[f"a{row}"].value,                   # "reference"
          manufacturer + " " + base_sheet[f"b{row}"].value.replace(manufacturer, ""), # "name"
          str(base_sheet[f"i{row}"].value),                   # "EAN13"
          base_sheet[f"c{row}"].value,                                     # "category"
          manufacturer + " " + base_sheet[f"b{row}"].value.replace(manufacturer, ""),                   # "Meta title"
          manufacturer + ","+ base_sheet[f"a{row}"].value  + ","+ base_sheet[f"b{row}"].value.replace(manufacturer, "").replace(" ", ",").replace("-", ",").replace("/", ","),                   # "Tags"
          manufacturer + ","+ base_sheet[f"a{row}"].value  + ","+ base_sheet[f"b{row}"].value.replace(manufacturer, "").replace(" ", ",").replace("-", ",").replace("/", ","),                   # "Keywords"
          slugify(manufacturer + "-"+ base_sheet[f"a{row}"].value),                   # "rewrite"
          base_sheet[f"e{row}"].value,
          base_sheet[f"f{row}"].value,
          base_sheet[f"g{row}"].value,
          base_sheet[f"h{row}"].value,
      ]
      import_sheet.append(temp_row)
    
wb.save(manufacturer.upper() + "_AMERICAN_ALL_STARS_IMPORT_CATALOGUE.xlsx")






