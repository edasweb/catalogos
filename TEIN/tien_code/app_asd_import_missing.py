

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s

def calculate_purcharse(group, price):
   percents = {
      "COILOVERS": 0.75,
      "SPRINGS": 0.7,
      "ABSORBERS": 0.60,}
   
   if percents.get(group):
      return price * Decimal(percents.get(group))
   return price * Decimal(percents.get("ABSORBERS"))
def discount(group):
   percents = {
      "COILOVERS": 15,
      "SPRINGS": 20,
      "ABSORBERS": 30,}
   
   if percents.get(group):
      return percents.get(group)
   return 30




office_sheet_name = "Folha1"
wb_office = openpyxl.load_workbook("from-dashboard.xlsx")

office_sheet_1 = wb_office[office_sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
saved = []
for row in range(1, office_sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    saved.append(office_sheet_1[f"c{row}"].value)


sheet_name = "Folha1"
wb = openpyxl.load_workbook("tein_all_we_need_as_base.xlsx")

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


sheet_no_need = wb["delete"]

no_need = []
for row in range(2, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    no_need.append(sheet_no_need[f"b{row}"].value)

    
no_need_count = 0
ranges = {}

already_in_use = {}
brand_name = None

for row in range(3, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    found_range = sheet_1[f"k{row}"].value if sheet_1[f"k{row}"].value != None else "ABSORBERS"
    if not sheet_1[f"d{row}"].value in no_need and already_in_use.get(sheet_1[f"d{row}"].value) !=  sheet_1[f"f{row}"].value:
      already_in_use[sheet_1[f"d{row}"].value] = sheet_1[f"f{row}"].value
      if not sheet_1[f"d{row}"].value in saved:
         print(sheet_1[f"d{row}"].value)
      
  






