
# generate bue one

import openpyxl
from datetime import date
from decimal import Decimal
sheet_name = "Folha1"
wb = openpyxl.load_workbook("tien_catalogue_new3.xlsx")
print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet



sheet_no_need = wb["delete"]

no_need = []
for row in range(2, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    no_need.append(sheet_no_need[f"b{row}"].value)

    
no_need_count = 0
brand_name = None
ranges = {}
for row in range(3, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    if brand_name != sheet_1[f"a{row}"].value:
        brand_name = sheet_1[f"a{row}"].value
        sheet_2.append([sheet_1[f"a{row}"].value, "", "", ""])
        sheet_2.append([
            "Part Number",
            "Model",
            "Chassis",
            "Item",
            "Price",
            "Note",
            "Year",
            "Drive System",
            "Displacement",
            "Range",
            "EDFC",
            "Sp Rate Ft (Kgf/mm)",
            "Sp Rate Rr (Kgf/mm)",
            "Std Ride Height Drop Ft (Mm)",
            "Std Ride Height Drop Rr(Mm)",
            "Recommended Ride Height Drop Max High Ft/mm",
            "Recommended Ride Height Drop Max Low Ft/mm",
            "Recommended Ride Height Drop Max High Rr/mm",
            "Recommended Ride Height Drop Max Low Rr/mm",
            "Matching Remarks"       
         ])
    print(sheet_1[f"g{row}"].value, " ", row)
    found_range = sheet_1[f"k{row}"].value if sheet_1[f"k{row}"].value != None else "ABSORBERS"
    if sheet_1[f"k{row}"].value == None:
        print("Row: ", row)
    
    
    print(sheet_1[f"f{row}"].value)
    if not sheet_1[f"d{row}"].value in no_need:
        temp_row = [
            sheet_1[f"d{row}"].value,  
            sheet_1[f"b{row}"].value, 
            sheet_1[f"c{row}"].value, 
            sheet_1[f"e{row}"].value, 
            Decimal(sheet_1[f"f{row}"].value) * Decimal(1.2), 
            sheet_1[f"g{row}"].value, 
            sheet_1[f"h{row}"].value, 
            sheet_1[f"i{row}"].value, 
            sheet_1[f"j{row}"].value, 
            found_range, 
            sheet_1[f"l{row}"].value, 
            sheet_1[f"m{row}"].value, 
            sheet_1[f"n{row}"].value, 
            sheet_1[f"o{row}"].value, 
            sheet_1[f"p{row}"].value, 
            sheet_1[f"q{row}"].value, 
            sheet_1[f"r{row}"].value, 
            sheet_1[f"s{row}"].value, 
            sheet_1[f"t{row}"].value, 
            sheet_1[f"u{row}"].value, 

            
        ]
        sheet_2.append(temp_row)
    else:
        no_need_count = no_need_count + 1
        

print("No need: ", no_need_count)
wb.save("tien_catalogue_new4.xlsx")


