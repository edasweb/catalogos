
# generate bue one

import openpyxl
from datetime import date
from decimal import Decimal
sheet_name = "Order"
wb = openpyxl.load_workbook("references_to_filtered.xlsx") 
print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


brand_name = None
for row in range(1, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    if brand_name != sheet_1[f"a{row}"].value:
        brand_name = sheet_1[f"a{row}"].value
        sheet_2.append([sheet_1[f"a{row}"].value, "", "", ""])
        sheet_2.append(["Part Number", "Car Application", "Description", "Retail Price Ex VAT"])
    print(sheet_1[f"g{row}"].value, " ", row)
    temp_row = [
        sheet_1[f"d{row}"].value, 
        sheet_1[f"a{row}"].value + " " + str(sheet_1[f"b{row}"].value)  + (" " + str(sheet_1[f"m{row}"].value) if str(sheet_1[f"m{row}"].value) != None else "") , 
        str(sheet_1[f"a{row}"].value) + " " + str(sheet_1[f"b{row}"].value)  + " TIEN " + (" " + str(sheet_1[f"c{row}"].value) if str(sheet_1[f"c{row}"].value) != None else "" ) + " " + str(sheet_1[f"e{row}"].value).upper() ,  
        (Decimal(sheet_1[f"g{row}"].value) * Decimal(1.2)) if row > 1 else sheet_1[f"g{row}"].value, 
    ]
    sheet_2.append(temp_row)
    


wb.save("tien_catalogue_new.xlsx")


