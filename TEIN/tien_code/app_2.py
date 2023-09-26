

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s

def calculate_purcharse(group, price):
   percents = {
      "COILOVERS": 0.75,
      "SPRINGS": 0.7,
      "ABSORBERS": 0.60,}
   
   if percents.get(group):
      return price * Decimal(percents.get(group))
   return price * Decimal(percents.get("ABSORBERS"))
def discount(group):
   percents = {
      "COILOVERS": 15,
      "SPRINGS": 20,
      "ABSORBERS": 30,}
   
   if percents.get(group):
      return percents.get(group)
   return 30



sheet_name = "Folha1"
wb = openpyxl.load_workbook("tien_catalogue_new3.xlsx")
print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


sheet_no_need = wb["delete"]

no_need = []
for row in range(2, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    no_need.append(sheet_no_need[f"b{row}"].value)

    
no_need_count = 0
ranges = {}


brand_name = None
sheet_2.append(["Manufacturer","Supplier","Price","Purchase","%","VAT","Reference","Name","EAN13","Category","Meta title","Tags","Keywords","Rewrite"])

for row in range(3, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    found_range = sheet_1[f"k{row}"].value if sheet_1[f"k{row}"].value != None else "ABSORBERS"
    if not sheet_1[f"d{row}"].value in no_need:
      temp_row = [
          "Tein",                                # "manufacturer"
          "Tein",                                # "Supplier"
          Decimal(sheet_1[f"f{row}"].value) * Decimal(1.2),                   # "price"
          calculate_purcharse(sheet_1[f"k{row}"].value, Decimal(sheet_1[f"f{row}"].value) * Decimal(1.2)), # "purchase"
          discount(sheet_1[f"k{row}"].value),                                       # " % "
          "7",                                        # "VAT"
          sheet_1[f"d{row}"].value,                   # "reference"
          sheet_1[f"a{row}"].value.title() + " " + str(sheet_1[f"b{row}"].value)  + " " + str(sheet_1[f"c{row}"].value)  +" TEIN ".title() + found_range.title() + " " + str(sheet_1[f"e{row}"].value).title(), # "name"
          "",                   # "EAN13"
          "Home",                                     # "category"
          sheet_1[f"a{row}"].value.title() + " " + str(sheet_1[f"b{row}"].value)  + " " + str(sheet_1[f"c{row}"].value)  +" TEIN ".title() + found_range.title() + " " + str(sheet_1[f"e{row}"].value).title(), # "Meta title"
          "Tein," + sheet_1[f"d{row}"].value + ","+ sheet_1[f"a{row}"].value.title() + ","+found_range.title() + "," + str(sheet_1[f"b{row}"].value).replace("-", ",").replace("/", ",") +"," + str(sheet_1[f"c{row}"].value) +","+ str(sheet_1[f"e{row}"].value).replace("-", ",").replace("/", ","), # "Meta Tags"
          "Tein," + sheet_1[f"d{row}"].value + ","+ sheet_1[f"a{row}"].value.title() + ","+found_range.title() + "," + str(sheet_1[f"b{row}"].value).replace("-", ",").replace("/", ",") +"," + str(sheet_1[f"c{row}"].value) +","+ str(sheet_1[f"e{row}"].value).replace("-", ",").replace("/", ","), # "Keywords"
          slugify("Tein-" + sheet_1[f"d{row}"].value),                   # "rewrite"
      ]
      sheet_2.append(temp_row)


  
    
wb.save("tien_all_stars_import_catalogue_new.xlsx")






