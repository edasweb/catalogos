

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s


sheet_name = "base+add"
wb = openpyxl.load_workbook("TIEN_RRP_FULL.xlsx") 
print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"Found", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
sheet_need = wb.create_sheet(title=f"Need", index=1) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


sheet_delete = wb["delete"]

delete_references = []
for row in range(2, sheet_delete.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    delete_references.append(str(sheet_delete[f'b{row}'].value))
  

brand_name = None


temp_row = ["Found"]
sheet_2.append(temp_row)
count = 0
for row in range(1, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    if sheet_1[f'd{row}'].value in delete_references:
       count = count + 1
       temp_row = [sheet_1[f'd{row}'].value]
       print(count, ": ", sheet_1[f'd{row}'].value)
       sheet_2.append(temp_row)
    else:
          
          temp_row = [
                   sheet_1[f'a{row}'].value,
                   sheet_1[f'b{row}'].value,
                   sheet_1[f'c{row}'].value,
                   sheet_1[f'd{row}'].value,
                   sheet_1[f'e{row}'].value,
                   sheet_1[f'f{row}'].value,
                   sheet_1[f'g{row}'].value,
                   sheet_1[f'h{row}'].value,
                   sheet_1[f'i{row}'].value,
                   sheet_1[f'j{row}'].value,
                   sheet_1[f'k{row}'].value,
                   sheet_1[f'l{row}'].value,
                   sheet_1[f'm{row}'].value,
                   sheet_1[f'n{row}'].value,
                   sheet_1[f'o{row}'].value,
                   sheet_1[f'p{row}'].value,
                   sheet_1[f'q{row}'].value,
                   sheet_1[f'r{row}'].value,
                   sheet_1[f's{row}'].value,
                   sheet_1[f't{row}'].value,
                   sheet_1[f'u{row}'].value,
                   sheet_1[f'v{row}'].value,
                   sheet_1[f'w{row}'].value,
                   sheet_1[f'x{row}'].value,
                   sheet_1[f'y{row}'].value,
                   sheet_1[f'z{row}'].value,
                   ]
          sheet_need.append(temp_row)
    
    
wb.save("references_to_filtered.xlsx")






