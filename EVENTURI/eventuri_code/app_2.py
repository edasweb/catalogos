

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s

sheet_name = "Global"
wb = openpyxl.load_workbook("WORK_FILE.xlsx") 
# print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


brand_name = None
sheet_2.append(["Manufacturer","Supplier","Price","Purchase","%","VAT","Reference","Name","EAN13","Category","Meta title","Tags","Keywords","Rewrite","Size"])
group = None
for row in range(2, 157): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    group = sheet_1[f"a{row}"].value if sheet_1[f"a{row}"].value != None else group
    temp_row = [
        "Eventuri",                                # "manufacturer"
        "Eventuri",                                # "Supplier"
        sheet_1[f"f{row}"].value if not str(sheet_1[f"f{row}"].value).startswith("=") else str(str(sheet_1[f"g{row}"].value/1.05)),                   # "price"
        Decimal(sheet_1[f"f{row}"].value if not str(sheet_1[f"f{row}"].value).startswith("=") else Decimal(str(sheet_1[f"g{row}"].value/1.05))) * Decimal(0.60), # "purchase"
        "15",                                       # " % "
        "7",                                        # "VAT"
        sheet_1[f"b{row}"].value,                   # "reference"
        sheet_1[f"c{row}"].value,                   # "product"
        "",                                         # "EAN13"
        "Home",                                     # "Category"
        sheet_1[f"c{row}"].value,                    # "Meta title"
        "Eventuri," + str(sheet_1[f"b{row}"].value) + ","+ ("" if sheet_1[f"c{row}"].value == None else sheet_1[f"c{row}"].value), # "Tags"
        "Eventuri," + str(sheet_1[f"b{row}"].value) + ","+ ("" if sheet_1[f"c{row}"].value == None else sheet_1[f"c{row}"].value), # "Keywords"
        slugify("Eventuri-" + sheet_1[f"b{row}"].value),                   # "rewrite"
        sheet_1[f"h{row}"].value,                   # "rewrite"
       
    ]
    sheet_2.append(temp_row)
    
wb.save("eventuri_all_stars_import_catalogue_new.xlsx")






