
# generate bue one

import openpyxl
from datetime import date

def get_value(firstValue, sheet, row, letter):
    firstValue = str(firstValue).replace("=", "")
    # print(sheet[firstValue.split("/")[0]].value)
    # print(sheet[firstValue.split("*")[1]].value)
    if "/" in str(sheet[f"{letter}{row}"]):
        values = sheet[f"{letter}{row}"].split("/")
        print(sheet[f"{letter}{row}"])
        print("row", row)
        print(float(values[1]))
        if str(values[1]) == "1.2" or  str(values[1]) == "1,2":
            return float(sheet[values[0]].value)/1.2
        # return float(sheet[values[0]].value)/float(sheet[values[1]].value) if not str(values[1]).isnumeric() else values[1]
        return f"Numberrr: {values[1]}"
    elif "*" in str(sheet[f"{letter}{row}"]):
        values = sheet[f"{letter}{row}"].split("*")
        return float(sheet[values[0]].value)*float(sheet[values[1]].value)
    elif "+" in str(sheet[f"{letter}{row}"]):
        values = sheet[f"{letter}{row}"].split("+")
        return float(sheet[values[0]].value)+float(sheet[values[1]].value)
    elif "-" in str(sheet[f"{letter}{row}"]):
        values = sheet[f"{letter}{row}"].split("-")
        return float(sheet[values[0]].value)-float(sheet[values[1]].value)

    

    # and str(sheet_1[f"e{row}"].value).contains("=") not in "*") else  float(sheet_1[str(sheet_1[f"e{row}"].value).replace("=","").split("/")[0]].value)/float(1)

sheet_name = "Global"
wb = openpyxl.load_workbook("WORK_FILE.xlsx") 
print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


brand_name = None
group = None
for row in range(1, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    if brand_name != sheet_1[f"l{row}"].value:
        brand_name = sheet_1[f"l{row}"].value
        if sheet_1[f"l{row}"].value != None:
            sheet_2.append([sheet_1[f"l{row}"].value.upper() if sheet_1[f"l{row}"] else "" , "", "", "", "", "", "", "", ""])
        sheet_2.append(["Aplication", "Part Number", "Description", "Filter Type","Retail Price Ex VAT £", "Retail Price Ex VAT €", "Retail Price", "Package size in cm",  "Brand"])
    # print(str(sheet_1[f"e{row}"].value).replace("=","").split("/")[0])
    print(row)
    
    
    group = sheet_1[f"a{row}"].value if sheet_1[f"a{row}"].value != None else group
    
    
    temp_row = [
        group, 
        sheet_1[f"b{row}"].value, 
        sheet_1[f"c{row}"].value, 
        sheet_1[f"d{row}"].value, 
        sheet_1[f"e{row}"].value if not str(sheet_1[f"e{row}"].value).startswith("=") else str(str(sheet_1[f"g{row}"].value/1.2)), 
        sheet_1[f"f{row}"].value, 
        sheet_1[f"g{row}"].value, 
        sheet_1[f"h{row}"].value, 
        sheet_1[f"l{row}"].value, 
    ]
    sheet_2.append(temp_row)
    


wb.save("eventuri_catalogue_new.xlsx")


