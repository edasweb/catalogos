

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s


sheet_name = "Global"
wb = openpyxl.load_workbook("WORK_FILE.xlsx") 
print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


brand_name = None


sheet_2.append(["Brand","Price Ex VAT £","Price Ex VAT €","Price Ex VAT $","Reference","name", "Package size in cm","Filter Type"])

for row in range(2, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    temp_row = [
        "Eventuri",                                # "manufacturer"
        sheet_1[f"e{row}"].value if not str(sheet_1[f"e{row}"].value).startswith("=") else str(str(sheet_1[f"g{row}"].value/1.2)), 
        sheet_1[f"f{row}"].value if not str(sheet_1[f"f{row}"].value).startswith("=") else str(str(sheet_1[f"g{row}"].value/1.05)), 
        sheet_1[f"g{row}"].value,
        sheet_1[f"b{row}"].value,
        sheet_1[f"c{row}"].value,
        sheet_1[f"h{row}"].value,
        sheet_1[f"d{row}"].value,
    ]
    sheet_2.append(temp_row)
    
wb.save("templace_csv_clientes_new.csv")






