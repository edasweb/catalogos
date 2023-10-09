

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s

sheet_name = "Global"
wb = openpyxl.load_workbook("WORK_FILE.xlsx") 
# print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


brand_name = None
sheet_2.append(["Manufacturer","Supplier","Price","Purchase","%","VAT","Reference","Name","Category","Meta title","Tags","Keywords","Rewrite","Size","Weight"])
group = ""
for row in range(2, 158): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    group = sheet_1[f"a{row}"].value if sheet_1[f"a{row}"].value != None else group
    group = "" if group == "OFF" else group
    temp_row = [
        "Eventuri",                                # "manufacturer"
        "Eventuri",                                # "Supplier"
        sheet_1[f"e{row}"].value,                   # "price"
        sheet_1[f"e{row}"].value * Decimal(0.60), # "purchase"
        "25",                                       # " % "
        "7",                                        # "VAT"
        sheet_1[f"b{row}"].value,                   # "reference"
        sheet_1[f"c{row}"].value,                   # "name"
        "Home",                                     # "Category"
        sheet_1[f"c{row}"].value,                    # "Meta title"
        "Eventuri," + str(sheet_1[f"b{row}"].value) + ","+ (sheet_1[f"c{row}"].value.replace(" ", ","))+ ","+ (group.replace(" ", ",")), # "Tags"
        "Eventuri," + str(sheet_1[f"b{row}"].value) + ","+ (sheet_1[f"c{row}"].value.replace(" ", ","))+ ","+ (group.replace(" ", ",")), # "Keywords"
        slugify("Eventuri-" + sheet_1[f"b{row}"].value),                   # "rewrite"
        sheet_1[f"f{row}"].value,                   # "size"
        sheet_1[f"g{row}"].value,                   # "weight"
       
    ]
    sheet_2.append(temp_row)
    
wb.save("eventuri_all_stars_import_catalogue_new.xlsx")






