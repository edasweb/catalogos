

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s

def getString(value):
    if value:
        return " "+str(value)
    
    return ""
def getSplitStart(value):
    return value[value.find("-")+2:].strip() if value.find("-") != -1 else value
    





sheet_name = "all_2"
wb = openpyxl.load_workbook("separated_new_update.xlsx") 
# print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"found- {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet




sheet_on_asd_web_name = "onasdwebsite"

sheet_asd_website = wb[sheet_on_asd_web_name] # To accsses the a sheet in the workbook. And create a sheet object.

reference_in_asd_web = list()

for row in range(1, sheet_asd_website.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    # print(sheet_asd_website[f"c{row}"].value)
    reference_in_asd_web.append(sheet_asd_website[f"c{row}"].value)








sheet_sizes_name = "sizes"

sheet_sizes = wb[sheet_sizes_name] # To accsses the a sheet in the workbook. And create a sheet object.

sizes = dict()

for row in range(2, sheet_sizes.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    sizes[sheet_sizes[f"a{row}"].value.strip()] = sheet_sizes[f"b{row}"].value/1000


all = list()

for row in range(2, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    all.append(sheet_1[f"i{row}"].value.strip())

# print(all)
sheet_their_exel_name = "their_exel"

sheet_their_exel = wb[sheet_their_exel_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2.append(["Manufacturer","Supplier","Price EUR","Purchase EUR","Price USD","Purchase USD","%","VAT","Reference","Name","Category","Meta title","Tags","Keywords","Rewrite", "Depth","Width","Height","Weight"
])


for row in range(2, sheet_their_exel.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    reference = sheet_their_exel[f"a{row}"].value.strip()
    if not "+" in sheet_their_exel[f"a{row}"].value  and not reference in all and sheet_their_exel[f"b{row}"].value != None and str(sheet_their_exel[f"b{row}"].value).isnumeric() and not reference in reference_in_asd_web and sheet_their_exel[f"h{row}"].value.upper() != "Coming Soon".upper():
        # print(sheet_their_exel[f"a{row}"].value.strip())
        # print(row)
        temp_row = list()
        # print(sheet_their_exel[f"b{row}"].value)
        temp_row = [
            "P3 Gauges",                                # "manufacturer"
            "P3 Gauges",                                # "Supplier"
            sheet_their_exel[f"b{row}"].value,                   # "price"
            Decimal(sheet_their_exel[f"b{row}"].value) * Decimal(0.58), # "purchase"
            sheet_their_exel[f"b{row}"].value,                   # "price"
            Decimal(sheet_their_exel[f"b{row}"].value) * Decimal(0.58), # "purchase"
            "35",                                       # " % "
            "7",                                        # "VAT"
            sheet_their_exel[f"a{row}"].value,                   # "reference"
            sheet_their_exel[f"c{row}"].value,  # "name"
            "Home",                   # "category"
            sheet_their_exel[f"c{row}"].value,    # "Meta title"
            "P3 Gauges," + sheet_their_exel[f"a{row}"].value + "," +sheet_their_exel[f"c{row}"].value,                   # "Tags"
            "P3 Gauges," + sheet_their_exel[f"a{row}"].value + "," +sheet_their_exel[f"c{row}"].value.replace(" ", ",").replace("/", ","),             # "Keywords"
            slugify("P3 Gauges-" + sheet_their_exel[f"a{row}"].value),                   # "rewrite"
            sheet_their_exel[f"d{row}"].value if sheet_their_exel[f"d{row}"].value != None else sheet_their_exel[f"d{row}"].value,                   # "deph"
            sheet_their_exel[f"e{row}"].value if sheet_their_exel[f"e{row}"].value != None else sheet_their_exel[f"d{row}"].value,                   # "width"
            sheet_their_exel[f"f{row}"].value if sheet_their_exel[f"f{row}"].value != None else sheet_their_exel[f"d{row}"].value,                   # "height"
            sizes[sheet_their_exel[f"g{row}"].value.strip()] if sheet_their_exel[f"g{row}"].value != None else sheet_their_exel[f"g{row}"].value,                   # "weight"
        
       
    ]
        sheet_2.append(temp_row)
    
wb.save("p3_found_not_comming_all_expect_asd_stars_import_catalogue_new.xlsx")


    






