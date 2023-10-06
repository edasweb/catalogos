

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s

def getString(value):
    if value:
        return " "+str(value)
    
    return ""
def getSplitStart(value):
    return value[value.find("-")+2:].strip() if value.find("-") != -1 else value





wb = openpyxl.load_workbook("separated_new_update.xlsx") 




sheet_name_on_base = "base"
# print(wb.sheetnames) 

sheet_on_base = wb[sheet_name_on_base] # To accsses the a sheet in the workbook. And create a sheet object.

all_refererences = list()
sheet_2 = wb.create_sheet(title=f"missing_member- {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet

for row in range(1, sheet_on_base.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    if "+" not in sheet_on_base[f"a{row}"].value:
        all_refererences.append(sheet_on_base[f"a{row}"].value.strip())


for ref in all_refererences: # To iterate over all the row and column of the sheet and get each value.

    if ref.startswith("R") and "L"+ref[1:] not in all_refererences:
        sheet_2.append([ref, "L"+ref[1:]])
    elif ref.startswith("L") and "R"+ref[1:] not in all_refererences:
        sheet_2.append([ref, "R"+ref[1:]])

    

wb.save("p3_missing_from_supplier.xlsx")






