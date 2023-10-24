

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
from pprint import pprint
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s


wb = openpyxl.load_workbook("websites_references.xlsx") 
# print(wb.sheetnames) 
sheet_name = "asm"
asm_sheet = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
asm_references = list()
found_sheet = wb.create_sheet(title=f"found - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
not_found_sheet = wb.create_sheet(title=f"not_found - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


matching_references = dict()
for row in range(1, asm_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    reference = asm_sheet[f"c{row}"].value
    if reference != None: 
      asm_references.append(reference.strip())
      matching_references[reference.strip()] = dict()





# pprint(matching_references)
sheet_name = "asd"
asd_sheet = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
asd_references = list()
import_matching_sheet = wb.create_sheet(title=f"import_matching - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet

for row in range(1, asd_sheet.max_row+1): # To iterate over all the row and column of the sheet and get each value.
    if asd_sheet[f"c{row}"].value  != None:
      asd_references.append(asd_sheet[f"c{row}"].value.strip())


found_references = []
for matching_reference in matching_references:
  if not matching_reference in asd_references:
      for reference in asd_references:
        lookup = matching_reference.upper().replace("-Z", "").replace("LX", "")
        if lookup.startswith("X"):
          lookup = lookup[1:]
        if lookup.startswith("B"):
          lookup = lookup[1:]
        if lookup.startswith("V"):
          lookup = lookup[1:]
          #  pprint(lookup)
        if lookup.startswith("UB"):
            lookup = lookup[2:]
        
        if lookup in reference:
            if matching_references[matching_reference].get("matching") == None:
              matching_references[matching_reference]["matching"] = lookup
              matching_references[matching_reference]["products"] = [reference]
            else:
              matching_references[matching_reference]["products"].append(reference)
  else:
     found_references.append(matching_reference)
     
  
  
sheet_name = "import_asd"
import_asd_sheet = wb[sheet_name]
used = {}
in_use = None
for matching_reference in matching_references:
    if matching_references[matching_reference].get("matching") != None:
      for reference in matching_references[matching_reference]["products"]:
           for row in range(1, import_asd_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
              if import_asd_sheet[f"i{row}"].value == reference and not import_asd_sheet[f"i{row}"].value.startswith("R"):
                temp_row = list()
                for column in range(1, import_asd_sheet.max_column + 1):
                    
                    temp_row.append(import_asd_sheet.cell(row, column).value)
                in_use = matching_reference.upper()
                if matching_reference.upper().startswith("LX"):
                   if "LA" in  import_asd_sheet[f"i{row}"].value.upper():
                      in_use = matching_reference.upper().replace("LX", "LA")
                   elif "L3" in  import_asd_sheet[f"i{row}"].value.upper():
                      in_use = matching_reference.upper().replace("LX", "L3")
                elif matching_reference.upper().startswith("X"):
                   if "LA" in  import_asd_sheet[f"i{row}"].value.upper():
                      in_use = "A" + matching_reference.upper()[1:]
                   elif "L3" in  import_asd_sheet[f"i{row}"].value.upper():
                      in_use = "3" + matching_reference.upper()[1:]
                elif matching_reference.upper().startswith("P3"):
                   if "LA" in  import_asd_sheet[f"i{row}"].value.upper():
                      in_use = "LA" + matching_reference.upper()
                   elif "L3" in  import_asd_sheet[f"i{row}"].value.upper():
                      in_use = "3" + matching_reference.upper()
                elif matching_reference.upper().startswith("3P"):
                   if "LA" in  import_asd_sheet[f"i{row}"].value.upper():
                      in_use = "LA" + matching_reference.upper()
                   elif "L3" in  import_asd_sheet[f"i{row}"].value.upper():
                      in_use = matching_reference.upper()
                elif matching_reference.upper().startswith("BP"):
                   if "LA" in  import_asd_sheet[f"i{row}"].value.upper() or  "UA" in  import_asd_sheet[f"i{row}"].value.upper():
                      in_use = matching_reference.upper()
                   elif "L3" in  import_asd_sheet[f"i{row}"].value.upper() or "U3" in  import_asd_sheet[f"i{row}"].value.upper():
                      in_use = "V" + matching_reference.upper()[1:]
                
                
                temp_row.append(in_use.replace("-Z", ""))
                # print(row)
                import_matching_sheet.append(temp_row)
    
    elif matching_references[matching_reference].get("matching") == None:
       if not matching_reference in asd_references:
          not_found_sheet.append([matching_reference])

for found_reference in found_references:
  for row in range(1, import_asd_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
      if import_asd_sheet[f"i{row}"].value == found_reference:
          temp_row = list()
          # print(row)

          for column in range(1, import_asd_sheet.max_column + 1):
              temp_row.append(import_asd_sheet.cell(row, column).value)
          temp_row.append(found_reference)
            # print(row)
          found_sheet.append(temp_row)
        

wb.save("asm_matching_sheet.xlsx")






