
# generate bue one

import openpyxl
from datetime import date



wb = openpyxl.load_workbook("complete_list.xlsx") 

work_sheet_name = "Worksheet3"

work_sheet = wb[work_sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_new = wb.create_sheet(title=f"news - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
sheet_update = wb.create_sheet(title=f"update - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
sheet_all = wb.create_sheet(title=f"all - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


for row in range(1, work_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = [
            work_sheet[f"a{row}"].value, 
            work_sheet[f"b{row}"].value, 
            work_sheet[f"c{row}"].value, 
            work_sheet[f"d{row}"].value, 
            work_sheet[f"e{row}"].value, 
            work_sheet[f"f{row}"].value, 
            work_sheet[f"g{row}"].value, 
            work_sheet[f"h{row}"].value, 
            work_sheet[f"i{row}"].value, 
            work_sheet[f"j{row}"].value, 
            work_sheet[f"k{row}"].value, 
            work_sheet[f"l{row}"].value, 
            work_sheet[f"m{row}"].value, 
            work_sheet[f"n{row}"].value,
            work_sheet[f"o{row}"].value,
    ]
    if work_sheet[f"n{row}"].value == "new":
        sheet_new.append(temp_row)
    else:
        sheet_update.append(temp_row)
    
    sheet_all.append(temp_row)

    


wb.save("separated_new_update.xlsx")


