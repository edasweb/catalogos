

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s

def getString(value):
    if value:
        return " "+str(value)
    
    return ""
def getSplitStart(value):
    return value[value.find("-")+2:].strip() if value.find("-") != -1 else value
    





sheet_name = "all"
wb = openpyxl.load_workbook("separated_new_update.xlsx") 
# print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name}_import- {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


sheet_sizes_name = "sizes"

sheet_sizes = wb[sheet_sizes_name] # To accsses the a sheet in the workbook. And create a sheet object.

sizes = dict()

for row in range(2, sheet_sizes.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    sizes[sheet_sizes[f"a{row}"].value.strip()] = sheet_sizes[f"b{row}"].value/1000

brand_name = None
sheet_2.append(["Manufacturer","Supplier","Price EUR","Purchase EUR","Price USD","Purchase USD","%","VAT","Reference","Name","Category","Meta title","Tags","Keywords","Rewrite", "Depth","Width","Height","Weight"
])

for row in range(2, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    temp_row = [
        "P3 Gauges",                                # "manufacturer"
        "P3 Gauges",                                # "Supplier"
        sheet_1[f"b{row}"].value,                   # "price"
        Decimal(sheet_1[f"b{row}"].value) * Decimal(0.58), # "purchase"
        sheet_1[f"b{row}"].value,                   # "price"
        Decimal(sheet_1[f"b{row}"].value) * Decimal(0.58), # "purchase"
        "35",                                       # " % "
        "7",                                        # "VAT"
        sheet_1[f"a{row}"].value,                   # "reference"
        sheet_1[f"c{row}"].value,  # "name"
        "Home",                   # "category"
        sheet_1[f"c{row}"].value,    # "Meta title"
        "P3 Gauges," + sheet_1[f"a{row}"].value + "," +sheet_1[f"c{row}"].value,                   # "Tags"
        "P3 Gauges," + sheet_1[f"a{row}"].value + "," +sheet_1[f"c{row}"].value.replace(" ", ",").replace("/", ","),             # "Keywords"
        slugify("P3 Gauges-" + sheet_1[f"a{row}"].value),                   # "rewrite"
        sheet_1[f"d{row}"].value.replace("'","").replace(" ","") if sheet_1[f"d{row}"].value != None else sheet_1[f"d{row}"].value,                   # "deph"
        sheet_1[f"e{row}"].value.replace("'","").replace(" ","") if sheet_1[f"e{row}"].value != None else sheet_1[f"d{row}"].value,                   # "width"
        sheet_1[f"f{row}"].value.replace("'","").replace(" ","") if sheet_1[f"f{row}"].value != None else sheet_1[f"d{row}"].value,                   # "height"
        sizes[sheet_1[f"g{row}"].value.strip()] if sheet_1[f"g{row}"].value != None else sheet_1[f"g{row}"].value,                   # "weight"
        
       
    ]
    sheet_2.append(temp_row)
    
wb.save("p3_all_all_stars_import_catalogue_new2.xlsx")






