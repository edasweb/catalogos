

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s

def getString(value):
    if value:
        return " "+str(value)
    
    return ""
def getSplitStart(value):
    return value[value.find("-")+2:].strip() if value.find("-") != -1 else value





wb = openpyxl.load_workbook("separated_new_update.xlsx") 






# print(already_in_website)


sheet_name = "all"

# print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name}_import- {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
on_date = list()
for row in range(2, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    reference = sheet_1[f"a{row}"].value
    if reference in on_date:
        print("Duplicated: ", reference)

    else:
        on_date.append(reference)



sheet_name_on_website = "on_website_2"
# print(wb.sheetnames) 

sheet_on_web_site = wb[sheet_name_on_website] # To accsses the a sheet in the workbook. And create a sheet object.

already_in_website = list()

for row in range(1, sheet_on_web_site.max_row + 1): # To iterate over all the row and column of the sheet and get each value.

    reference = sheet_on_web_site[f"d{row}"].value if sheet_on_web_site[f"d{row}"].value != None else None,
    # print("reference: ",type(reference[0]))
    if reference != None and reference[0] != None:
        if not reference[0] in on_date:
            print("Outdate:", reference[0])
            sheet_2.append([reference[0]])      

        


        
    
wb.save("p3_out_date_news_all_stars_import_catalogue_new2.xlsx")






