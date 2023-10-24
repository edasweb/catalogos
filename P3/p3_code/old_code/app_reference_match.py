
# generate bue one

import openpyxl
from datetime import date


sheet_name = "Worksheet"
wb = openpyxl.load_workbook("p3_work_file.xlsx") 
 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


def getString(value):
    if value:
        return str(value)
    
    return ""

brand_name = None
group = None
for row in range(1, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    if not "+" in sheet_1[f"a{row}"].value and sheet_1[f"e{row}"].value != "Coming Soon":
       
        temp_row = [
            sheet_1[f"a{row}"].value + " ", 
            sheet_1[f"b{row}"].value, 
            sheet_1[f"c{row}"].value, 
            sheet_1[f"d{row}"].value, 
            sheet_1[f"e{row}"].value, 
            sheet_1[f"f{row}"].value, 
            sheet_1[f"g{row}"].value, 
            sheet_1[f"h{row}"].value, 
            sheet_1[f"i{row}"].value, 
            sheet_1[f"j{row}"].value, 
            sheet_1[f"k{row}"].value, 
            sheet_1[f"l{row}"].value, 
        ]
        sheet_2.append(temp_row)
    


wb.save("work_file_new.xlsx")


