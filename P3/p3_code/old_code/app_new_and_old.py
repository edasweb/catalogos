

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s




sheet_name = "Worksheet"
wb = openpyxl.load_workbook("WORK_FILE.xlsx") 
print(wb.sheetnames) 

work_sheet = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
sheet_online = wb["online"] # To accsses the a sheet in the workbook. And create a sheet object.

onlines = []
for row in range(1, sheet_online.max_row + 1): 
      onlines.append(sheet_online[f"a{row}"].value.replace(" ", ""))
      

sheet_new = wb.create_sheet(title=f"New {sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
sheet_expired = wb.create_sheet(title=f"Discontinued {sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet

news = []
discontinued = []
all_references = []
for row in range(2, work_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    if(row<=887):
      all_references.append(work_sheet[f"e{row}"].value.replace(" ", ""))
      if not work_sheet[f"e{row}"].value.replace(" ", "") in onlines:
          sheet_new.append([work_sheet[f"e{row}"].value.replace(" ", "")])
          news.append(work_sheet[f"e{row}"].value.replace(" ", ""))
          
count = 0      
for online_reference in onlines: 
      count = count + 1
      if not online_reference in all_references:
          sheet_expired.append([online_reference])
          discontinued.append(online_reference)
      else:
          # print(online_reference, " in line ",  count)
          pass
        
print(news) 
print(discontinued) 
wb.save("expired_and_new.xlsx")






