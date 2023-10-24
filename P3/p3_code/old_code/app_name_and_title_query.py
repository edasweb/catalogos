

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s

def getString(value):
    if value:
        return " "+str(value)
    
    return ""


sheet_name = "new_scorpion_all_stars_import_c"
wb = openpyxl.load_workbook("name_and_title.xlsx") 
# print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


brand_name = None
sheet_2.append(["Reference","Name","Meta title","Query"])
all_query_toguether = ""
for row in range(2, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    query =  f"""UPDATE psnz_product_lang  INNER JOIN psnz_product ON psnz_product_lang.id_product = psnz_product.id_product SET psnz_product_lang.meta_title = "{sheet_1[f"c{row}"].value.strip().replace("  ", " ")}", psnz_product_lang.name = "{sheet_1[f"b{row}"].value.strip().replace("  ", " ")}" WHERE psnz_product.reference = "{sheet_1[f"a{row}"].value}"; """
    temp_row = [
                                       
        sheet_1[f"a{row}"].value,                                
        sheet_1[f"b{row}"].value,                   
        sheet_1[f"c{row}"].value,                   
        query,                    
    ]
    all_query_toguether = all_query_toguether + query
    # print(query)
    sheet_2.append(temp_row)
# sheet_2.append([all_query_toguether])
wb.save("querys.xlsx")

# print(all_query_toguether)






