

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s

def getString(value):
    if value:
        return " "+str(value)
    
    return ""

scropion_sheet_name = "Worksheet"
scropion_wb = openpyxl.load_workbook("Work_File.xlsx") 

# print(wb.sheetnames) 

scropion_sheet_1 = scropion_wb[scropion_sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

scropion_sheet_2 = scropion_wb.create_sheet(title=f"{scropion_sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


scropion_current_price = dict()
for row in range(2, scropion_sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    if(scropion_current_price.get(scropion_sheet_1[f"e{row}"].value)):
        print("Doubled SCP: ", scropion_sheet_1[f"e{row}"].value)
    else:
        details = {"purcharse":  Decimal(scropion_sheet_1[f"h{row}"].value) * Decimal(0.50), "price": scropion_sheet_1[f"h{row}"].value}
        scropion_current_price[scropion_sheet_1[f"e{row}"].value] = details



all_stars_sheet_name = "ASDScorpion"
all_stars_wb = openpyxl.load_workbook("ASDScorpion.xlsx") 

# print(wb.sheetnames) 

all_stars_sheet_1 = all_stars_wb[all_stars_sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

all_stars_sheet_2 = all_stars_wb.create_sheet(title=f"{all_stars_sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


all_stars_current_price = dict()

for row in range(2, all_stars_sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    if(all_stars_current_price.get(all_stars_sheet_1[f"c{row}"].value)):
        print("Doubled ASD: ", all_stars_sheet_1[f"c{row}"].value)
    else:
        print(row)
        details = {"purcharse":  Decimal(all_stars_sheet_1[f"b{row}"].value) * Decimal(0.50), "price": all_stars_sheet_1[f"b{row}"].value}
        all_stars_current_price[all_stars_sheet_1[f"c{row}"].value] = details

# print(scropion_current_price)
# print(all_stars_current_price)
news = list()

temp_row = [
            "Reference",
            "ASD Price",
            "SCR Price",
            "ASD Purcharse",
            "SCR Purcharse",]
scropion_sheet_2.append(temp_row)
for reference in scropion_current_price:
    if all_stars_current_price.get(reference):
        if scropion_current_price[reference]["price"] != all_stars_current_price[reference]["price"] or scropion_current_price[reference]["purcharse"] != all_stars_current_price[reference]["purcharse"]:
            print(reference, scropion_current_price[reference], " and ", all_stars_current_price[reference])
            temp_row = [
            reference,
            all_stars_current_price[reference]["price"],
            scropion_current_price[reference]["price"],
            all_stars_current_price[reference]["purcharse"],
            scropion_current_price[reference]["purcharse"],
        ]
            scropion_sheet_2.append(temp_row)
    else:
        news.append(reference)

        
    

    # temp_row = [
    #     "Scorpion",                                # "manufacturer"
    #     scropion_sheet_1[f"b{row}"].value,                                # "Supplier"
    #     scropion_sheet_1[f"h{row}"].value,                   # "price"
    #     Decimal(scropion_sheet_1[f"h{row}"].value) * Decimal(0.50), # "purchase"
    #     "35",                                       # " % "
    #     "7",                                        # "VAT"
    #     scropion_sheet_1[f"e{row}"].value,                   # "reference"
    #     scropion_sheet_1[f"b{row}"].value + " " + scropion_sheet_1[f"c{row}"].value  + getString(scropion_sheet_1[f"j{row}"].value) + getString(scropion_sheet_1[f"d{row}"].value)+ " "+ getString(scropion_sheet_1[f"l{row}"].value) + " "+ (" PD(" if scropion_sheet_1[f"k{row}"].value != None else "") + getString(scropion_sheet_1[f"k{row}"].value)+ (")" if scropion_sheet_1[f"k{row}"].value != None else "")   + " "+ (" TD(" if scropion_sheet_1[f"m{row}"].value != None else "") + getString(scropion_sheet_1[f"m{row}"].value) + (")" if scropion_sheet_1[f"m{row}"].value != None else ""),  # "product"
    #     "Home",                   # "category"
    #     scropion_sheet_1[f"b{row}"].value + " " + scropion_sheet_1[f"c{row}"].value + getString(scropion_sheet_1[f"j{row}"].value)  + getString(scropion_sheet_1[f"d{row}"].value)+ ""+ getString(scropion_sheet_1[f"l{row}"].value) + "" + getString(scropion_sheet_1[f"k{row}"].value)  + "" + getString(scropion_sheet_1[f"m{row}"].value) ,                   # "Meta title"
    #     "Scorpion," + scropion_sheet_1[f"b{row}"].value + "," + scropion_sheet_1[f"c{row}"].value + "," + getString(scropion_sheet_1[f"j{row}"].value) + "," + getString(scropion_sheet_1[f"d{row}"].value)+ ","+ getString(scropion_sheet_1[f"l{row}"].value) + ","+ getString(scropion_sheet_1[f"k{row}"].value)+  "," + getString(scropion_sheet_1[f"m{row}"].value),                   # "Tags"
    #     "Scorpion," + scropion_sheet_1[f"b{row}"].value.replace(" ", ",") + "," + scropion_sheet_1[f"c{row}"].value.replace(" ", ",") + "," + getString(scropion_sheet_1[f"j{row}"].value).replace(" ", ",") + "," + scropion_sheet_1[f"d{row}"].value.replace(" ", ",")+ ","+ getString(scropion_sheet_1[f"l{row}"].value).replace(" ", ",") + ","+ getString(scropion_sheet_1[f"k{row}"].value).replace(" ", ",")+  "," + getString(scropion_sheet_1[f"m{row}"].value).replace(" ", ",")  ,             # "Keywords"
    #     slugify("Scorpion-" + scropion_sheet_1[f"e{row}"].value),                   # "rewrite"
       
    # ]
    # scropion_sheet_2.append(temp_row)
    
scropion_wb.save("scorpion_prices.xlsx")






