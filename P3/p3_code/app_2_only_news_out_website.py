

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s

def getString(value):
    if value:
        return " "+str(value)
    
    return ""
def getSplitStart(value):
    return value[value.find("-")+2:].strip() if value.find("-") != -1 else value





wb = openpyxl.load_workbook("separated_new_update.xlsx") 




sheet_name_on_website = "on_website_2"
# print(wb.sheetnames) 

sheet_on_web_site = wb[sheet_name_on_website] # To accsses the a sheet in the workbook. And create a sheet object.

already_in_website = list()

for row in range(1, sheet_on_web_site.max_row + 1): # To iterate over all the row and column of the sheet and get each value.

    reference = sheet_on_web_site[f"d{row}"].value if sheet_on_web_site[f"d{row}"].value != None else None,
    # print("reference: ",type(reference[0]))
    if reference != None and reference[0] != None:
        if reference[0] in already_in_website:
            print("Duplicate:", reference[0])
        else:
            already_in_website.append(reference[0])

# print(already_in_website)


sheet_name = "news"

# print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name}_import- {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


brand_name = None
sheet_2.append(["Manufacturer","Supplier","Price EUR","Purchase EUR","Price USD","Purchase USD","%","VAT","Reference","Name","Category","Meta title","Tags","Keywords","Rewrite", "Depth","Width","Height","Weight", "Name 2", "Meta Title 2"
])
in_website_count = 1
for row in range(2, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    if not sheet_1[f"a{row}"].value in already_in_website:
        temp_row = [
            "P3 Gauges",                                # "manufacturer"
            "P3 Gauges",                                # "Supplier"
            sheet_1[f"b{row}"].value,                   # "price"
            Decimal(sheet_1[f"b{row}"].value) * Decimal(0.58), # "purchase"
            sheet_1[f"b{row}"].value,                   # "price"
            Decimal(sheet_1[f"b{row}"].value) * Decimal(0.58), # "purchase"
            "35",                                       # " % "
            "7",                                        # "VAT"
            sheet_1[f"a{row}"].value,                   # "reference"
            sheet_1[f"d{row}"].value,  # "name"
            "Home",                   # "category"
            sheet_1[f"d{row}"].value,    # "Meta title"
            "P3 Gauges," + sheet_1[f"a{row}"].value + "," +sheet_1[f"d{row}"].value,                   # "Tags"
            "P3 Gauges," + sheet_1[f"a{row}"].value + "," +sheet_1[f"d{row}"].value.replace(" ", ",").replace("/", ","),             # "Keywords"
            slugify("P3 Gauges-" + sheet_1[f"a{row}"].value),                   # "rewrite"
            sheet_1[f"i{row}"].value,                   # "reference"
            sheet_1[f"j{row}"].value,                   # "reference"
            sheet_1[f"k{row}"].value,                   # "reference"
            sheet_1[f"l{row}"].value,                   # "reference"
            getSplitStart(sheet_1[f"d{row}"].value),  # "name"
            getSplitStart(sheet_1[f"d{row}"].value),    # "Meta title"
        
        ]
        sheet_2.append(temp_row)
    else:
        print(in_website_count, " - already in webiste: ", sheet_1[f"a{row}"].value)
        in_website_count = in_website_count + 1
wb.save("p3_only_news_all_stars_import_catalogue_new2.xlsx")






