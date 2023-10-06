
# generate bue one

import openpyxl
from datetime import date



brands = ["VW", "BMW", "Audi", "MINI", "Mitsubish", "Honda","Hyundai","Mercedes","Chevrolet","Cadillac","Buick","Ford","Toyota","Subaru","SEAT","All Models", "Mazda", "Porsche", "Fiat", "Dodge"]

def find_brand(name):
    found = "Check It"
    for brand in brands:
        if brand in name:
            found =  brand
            break
        
    return found



wb = openpyxl.load_workbook("work_file_new.xlsx") 

asd_sheet_name = "ASDList"

 

asd_sheet_1 = wb[asd_sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.



asd_references = list()
for row in range(2, asd_sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    asd_references.append(asd_sheet_1[f"c{row}"].value)
    
   



p3_sheet_name = "Worksheet2"


p3_sheet_1 = wb[p3_sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{p3_sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet



par_saved = list()
already_saved = list()
for row in range(1, p3_sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    last_value = ""
    matching_reference = ""

    if p3_sheet_1[f"a{row}"].value in asd_references:
        matching_reference = p3_sheet_1[f"a{row}"].value
    else:
        first_letter = p3_sheet_1[f"a{row}"].value[0]
        other_letters = p3_sheet_1[f"a{row}"].value[1:]
        reference = p3_sheet_1[f"a{row}"].value
        if first_letter in "LR":
            if other_letters in asd_references:
                matching_reference = other_letters
                if f"L"+other_letters in par_saved or "R"+other_letters in par_saved:
                    last_value = "new"
                    matching_reference = ""
                else:
                    par_saved.append(reference)
                
            else:
                par_saved.append(reference)
                last_value = "new"
                matching_reference = ""
        else:
            par_saved.append(reference)
            last_value = "new"
            matching_reference = ""
    
    
    if not p3_sheet_1[f"a{row}"].value in already_saved:
    
        temp_row = [
            p3_sheet_1[f"a{row}"].value, 
            p3_sheet_1[f"b{row}"].value, 
            p3_sheet_1[f"c{row}"].value, 
            p3_sheet_1[f"d{row}"].value, 
            p3_sheet_1[f"e{row}"].value, 
            p3_sheet_1[f"f{row}"].value, 
            p3_sheet_1[f"g{row}"].value, 
            p3_sheet_1[f"h{row}"].value, 
            p3_sheet_1[f"i{row}"].value, 
            p3_sheet_1[f"j{row}"].value, 
            p3_sheet_1[f"k{row}"].value, 
            p3_sheet_1[f"l{row}"].value, 
            matching_reference, 
            last_value, 
            find_brand(p3_sheet_1[f"d{row}"].value), 
        ]
        already_saved.append(p3_sheet_1[f"a{row}"].value)
        sheet_2.append(temp_row)
    


wb.save("complete_list.xlsx")


