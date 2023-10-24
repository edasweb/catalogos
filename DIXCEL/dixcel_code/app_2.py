

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
from pprint import pprint
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s


wb = openpyxl.load_workbook("start.xlsx") 
# print(wb.sheetnames) 
sheet_name = "all"
all_sheet = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

all_new_sheet = wb.create_sheet(title=f"all-new - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
error_sheet = wb.create_sheet(title=f"error - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
types_sheet = wb.create_sheet(title=f"types - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
cars_sheet = wb.create_sheet(title=f"cars - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet

def getIndexStart(name, ref):
  key = None
  if "FR." in name:
    key = "FR."
  elif "FR " in name:
    key = "FR  "
  elif "FR-" in name:
    key = "FR-"
  elif ("FR" + "．") in name:
    key = ("FR" + "．")
  elif "RR." in name: 
     key = "RR."
  elif "RR " in name: 
     key = "RR "
  elif "RR-" in name: 
     key = "RR-"
  elif("RR" + "．") in name:
     key = ("RR" + "．")
    
  if key != None:
    return name[name.find(key)+3:].strip()
  
  
  # print(name, ref)
  
  return None


def check_alpha(word):
  
  word = word.strip()
  
  if word != "":
    if 'a' <= word[len(word)-1] <= "z" or 'A' <= word[len(word)-1] <='Z':
        return True
  
  return False

def getcarname(car):
    car4 = car3 = car2 = car1 = car
    cars = set()
    car = car.split(" ")
       
    if len(car) > 0:
       car1 = car[0]

       
    if "0" in car1 or "1" in car1 or "2" in car1 or "3" in car1 or "4" in car1 or "5" in car1 or "6" in car1 or "7" in car1 or "8" in car1 or "9" in car1 or not check_alpha(car1):
      pass
    else:
     cars.add(car1)

    if len(car) > 1:
       car2 = car[1]

    if "0" in car2 or "1" in car2 or "2" in car2 or "3" in car2 or "4" in car2 or "5" in car2 or "6" in car2 or "7" in car2 or "8" in car2 or "9" in car2 or not check_alpha(car2):
      pass
    else:
      cars.add(car2)

    if len(car) > 2:
       car3 = car[2]

    if "0" in car3 or "1" in car3 or "2" in car3 or "3" in car3 or "4" in car3 or "5" in car3 or "6" in car3 or "7" in car3 or "8" in car3 or "9" in car3 or not check_alpha(car3):
      pass
    else:
      cars.add(car3)

    if len(car) > 3:
       car4 = car[3]

    if "0" in car4 or "1" in car4 or "2" in car4 or "3" in car4 or "4" in car4 or "5" in car4 or "6" in car4 or "7" in car4 or "8" in car4 or "9" in car4 or not check_alpha(car4):
      pass
    else:
      cars.add(car4)

    return cars

def replace_no_need(name, type):
  name = name.replace(type + " ", "").replace("FR.", "").replace("FR ", "").replace("FR-", "").replace("FR.", "").replace(("FR" + "．"), "")
  name = name.replace(type + " ", "").replace("RR.", "").replace("RR ", "").replace("RR-", "").replace("RR.", "").replace(("RR" + "．"), "")
  
  return name.replace("TYPE","").replace("  ", " ")

types = set()
cars = set()
all_new_sheet.append(["Old Reference", "Supplier", "Part Nmber (new)", "Type", "Position", "Name", "Full Name", "Price", "EAN"])
types_sheet.append(["Gama", "Type"])
error_sheet.append(["Error", "Reference"])
cars_sheet.append(["Retrive the cars"])
for row in range(2, all_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    part_number = reference = all_sheet[f"a{row}"].value.replace(" ", "")
    
    if not "-" in reference:
      part_number = reference[0:2] + "-" + reference[3:]
    
    
    type = all_sheet[f"b{row}"].value.split(" ")
    if len(type) == 3 or len(type) == 2:
      type = type[1]
      if type == "RGM" or type == "RGX" or type == "RGS":
        type = "SHOE " + type
      types.add((type, "PADS" if "PAD" in all_sheet[f"b{row}"].value else ("DISCS" if "ROTOR" in all_sheet[f"b{row}"].value else "SHOE PADS")))
    else:
      print("Type not found in disc at row: ",row)

    
    name = position = all_sheet[f"c{row}"].value.upper()
    
    if "FR." in position or "FR " in position or "FR-" in position or ("FR" + "．") in position:
      position = "FRONT".upper()
    elif "RR." in position or "RR " in position or "RR-" in position or ("RR" + "．") in position:
      position = "REAR".upper()
    else:
      print("Position not found in disc for: ", reference)

    if position == "REAR" or position == "FRONT":
       car = name = getIndexStart(name, reference).strip().replace("  ", " ")
       cars = cars.union(getcarname(car))
    else:
      error_sheet.append(["Name not updated in disc for: " + reference, reference])
      

    price = Decimal(all_sheet[f"e{row}"].value /140)
    code = all_sheet[f"d{row}"].value
    supplier = "Dixcel".upper()
    position = position if position == "REAR" or position == "FRONT" else ""
    part = "PADS" if "PAD" in all_sheet[f"b{row}"].value else ("DISCS" if "ROTOR" in all_sheet[f"b{row}"].value else "SHOE PADS")
    if not "PAD" in all_sheet[f"b{row}"].value and not "ROTOR" in all_sheet[f"b{row}"].value and part != "SHOE PADS":
      error_sheet.append(["Type not defined for " + reference, reference])
    # pprint(type)
    
    full_name = f"{supplier} {type} type {position} {part} for {replace_no_need(name.upper(), type)}".upper().replace(code if code != None else "", "").replace("  ", " ")

    all_new_sheet.append([reference, supplier, part_number, type, position, name, full_name, price, code])

for type in types:
  types_sheet.append([type[0], type[1]])

for car in sorted(cars):
  cars_sheet.append([car])


    
    





        

wb.save("all_from_app.xlsx")






