

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s
from pprint import pprint


def read_excel_file(file, obj):


    wb = openpyxl.load_workbook("WORKFILE.xlsx") 
    
    # print(wb.sheetnames) 
    


    new_sheet = wb.create_sheet(title=f"new - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
    update_price_sheet = wb.create_sheet(title=f"update_price - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
    
    update_online_price_sheet = wb.create_sheet(title=f"update_price_online - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
    import_sheet = wb.create_sheet(title=f"import - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
    new_sheet.append(["Manufacturer","Supplier","Name","Category","Price","VAT","Purchase","Meta title","EAN13","Discount","Tags","Meta Keywords","Reference","Rewrite"])

    client_sheet = wb.create_sheet(title=f"client - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
    client_sheet.append(["Manufacturer","Name","Category","Price","EAN13","Reference"])

    asd_sheet = wb.create_sheet(title=f"asd - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
    out_date_sheet = wb.create_sheet(title=f"out_date - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


    sheet_name = "WEBSITE"

    website_sheet = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
    website_references = list()

    for row in range(1, website_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
        if website_sheet[f"a{row}"].value != None and website_sheet[f"b{row}"].value != 0 and website_sheet[f"c{row}"].value: 
            website_references.append(website_sheet[f"a{row}"].value.upper().strip())


    sheet_name = "WORKFILE"
    work_sheet = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
    work_references = list()
    
    for row in range(1, work_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
        if work_sheet[f"a{row}"].value != None:
            work_references.append(work_sheet[f"a{row}"].value.upper().strip() )



    for reference in website_references:
        reference = reference.strip()
        if not reference in work_references:
            out_date_sheet.append([reference])


    for row in range(1, 2): # To iterate over all the row and column of the sheet and get each value.
        other_columns = []
        print(work_sheet.max_column)
        for column in range(14, work_sheet.max_column + 1):
            other_columns.append(work_sheet.cell(row,column).value)

        import_sheet.append(["Manufacturer","Supplier","Name","Category","Price","VAT","Purchase","Meta title","EAN13","Discount","Tags","Meta Keywords","Reference","Rewrite"]+other_columns)

    for row in range(2, work_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.

        currencys = None
        if obj.currency != None:
            currencys = Currency.objects.filter(id=obj.currency.id)
        
        exchange_id = 1
        currency = None
        exchanges = {}
        if currencys:
            currency = currencys.first()
            exchange_id = currencys.first().id
            exchanges = {
            f"{currency.id}":{"column":f"whosales_price_{currency.title.lower()}","rate":currency.rate},
            # 3:{"column":"whosales_price_gbp","rate":Decimal(0.8)},
            # 4:{"column":"whosales_price_yan","rate":Decimal(7.71)},
            # 5:{"column":"whosales_price_yen","rate":Decimal(157,81)},

        }
        price_list_columns = ["reference","discount","price","whosales_price_euro"]
        if exchange_id != 1:
            price_list_columns.append(exchanges[exchange_id]["column"])
        update_price_sheet.append(price_list_columns)
        update_online_price_sheet.append(price_list_columns)
        

        exchange_rate = exchanges.get(exchange_id)["rate"] if exchanges.get(exchange_id) != None else 1
        brand = None
        # General
        reference = work_sheet[f"a{row}"].value.strip()
        discount = work_sheet[f"c{row}"].value
        discount_client = work_sheet[f"d{row}"].value
       
        price = Decimal(work_sheet[f"b{row}"].value) *  Decimal(exchange_rate) 


        # Price List and News
        who_sales_price_euro = Decimal(work_sheet[f"b{row}"].value) * exchange_rate * Decimal((100-discount)/100)
        who_sales_price_original = Decimal(work_sheet[f"b{row}"].value) * Decimal((100-discount)/100)




         # Import Catalogue
        supplier = "ARMASPEED"
        price = price
        purchase = who_sales_price_euro
        discount = discount
        vat = 7
        reference = reference
        name = supplier + " " + (work_sheet[f"e{row}"].value).replace(supplier, "")
        ean = work_sheet[f"f{row}"].value
        category = work_sheet[f"g{row}"].value if work_sheet[f"g{row}"].value != None else "Home"
        meta_title = name
        tags = supplier + "," + reference+ "," + name.replace(" ", ",").replace("-", ",").replace("/", ",")
        keywords = tags
        url_rewriten = tags
        weight = work_sheet[f"h{row}"].value
        size = work_sheet[f"i{row}"].value
        width = work_sheet[f"j{row}"].value
        height = work_sheet[f"k{row}"].value
        depth = work_sheet[f"l{row}"].value



        short_descrition = work_sheet[f"m{row}"].value

        temp_row = [
            supplier,
            supplier,
            name,
            category,
            price,
            vat,
            purchase,
            meta_title,
            ean,
            discount_client,
            tags,
            keywords,
            reference,
            url_rewriten
        ]
        if short_descrition != None:
            temp_row.append(short_descrition)

            

        if size != None:
            size =str(size).upper().split("X")
            width = size[0].strip()
            height = size[1].strip()
            depth = size[2].strip()
            temp_row.append(width)
            temp_row.append(height)
            temp_row.append(depth)
        elif width != None or height != None or depth != None:
            temp_row.append(width)
            temp_row.append(height)
            temp_row.append(depth)
        
        if weight != None:
            temp_row.append(weight)
        if short_descrition != None:
            temp_row.append(short_descrition)

        
        


        other_columns = []
        print(work_sheet.max_column)
        for column in range(14, work_sheet.max_column + 1):
            other_columns.append(work_sheet.cell(row,column).value)

        pprint(other_columns)
        import_sheet.append(temp_row+other_columns)

        temp_row = list()


        exchange_id = 1
        currency = None
        exchanges = {}
        if currencys:
            currency = currencys.first()
            exchange_id = currencys.first().id
            exchanges = {
            f"{currency.id}":{"column":f"whosales_price_{currency.title.lower()}","rate":currency.rate},
            # 3:{"column":"whosales_price_gbp","rate":Decimal(0.8)},
            # 4:{"column":"whosales_price_yan","rate":Decimal(7.71)},
            # 5:{"column":"whosales_price_yen","rate":Decimal(157,81)},

        }
            
        price_list_columns = ["reference","discount","price","whosales_price_euro"]

        if exchange_id != 1:
            price_list_columns.append(exchanges[exchange_id]["column"])









        wb.save(file.path)
        obj.update_file = file
        obj.save()
        return 1
        temp_row = [reference, discount_client, price, who_sales_price_euro]


        if exchange_id != 1:
            temp_row.append(who_sales_price_original) 

        update_price_sheet.append(temp_row)

        if reference in website_references:
            update_online_price_sheet.append(temp_row)

        supplier = obj.name
        price = price
        purchase = who_sales_price_euro
        discount = discount
        vat = 7
        reference = reference
        name = supplier + " " + (work_sheet[f"e{row}"].value).replace(supplier, "")
        ean = work_sheet[f"f{row}"].value
        category = work_sheet[f"g{row}"].value if work_sheet[f"g{row}"].value != None else "Home"
        meta_title = name
        tags = supplier + "," + reference+ "," + name.replace(" ", ",").replace("-", ",").replace("/", ",")
        keywords = tags
        url_rewriten = tags
        weight = work_sheet[f"h{row}"].value
        size = work_sheet[f"i{row}"].value
        width = work_sheet[f"j{row}"].value
        height = work_sheet[f"k{row}"].value
        depth = work_sheet[f"l{row}"].value



        short_descrition = work_sheet[f"m{row}"].value





        if work_sheet[f"a{row}"].value != None:
            work_references.append(work_sheet[f"a{row}"].value.strip())
            reference = work_sheet[f"a{row}"].value.strip()
            if not reference in website_references:
                temp_row = list()
                for col in range(1, work_sheet.max_column + 1):
                    # print(row, col)
                    cell_value = work_sheet.cell(row, col).value
                    temp_row.append(cell_value)
                    
                new_sheet.append(temp_row)
        currencys = None
        if obj.currency != None:
            currencys = Currency.objects.filter(id=obj.currency.id)
        
        exchange_id = 1
        currency = None
        exchanges = {}
        if currencys:
            currency = currencys.first()
            exchange_id = currencys.first().id
            exchanges = {
            f"{currency.id}":{"column":f"whosales_price_{currency.title.lower()}","rate":currency.rate},
            # 3:{"column":"whosales_price_gbp","rate":Decimal(0.8)},
            # 4:{"column":"whosales_price_yan","rate":Decimal(7.71)},
            # 5:{"column":"whosales_price_yen","rate":Decimal(157,81)},

        }
        price_list_columns = ["reference","discount","price","whosales_price_euro"]
        # brands = ["VW", "Acura", "Volkswagen", "BMW", "Audi", "MINI", "Mitsubish", "Honda","Hyundai","Mercedes","Chevrolet","Cadillac","Buick","Ford",
        #       "Toyota","Subaru","SEAT","All Models", "Mazda", "Porsche", "Fiat", "Dodge", "jeep", "Suzuki"]


        rest = []
        weight = work_sheet[f"h{row}"].value
        size = work_sheet[f"i{row}"].value
        width = work_sheet[f"j{row}"].value
        height = work_sheet[f"k{row}"].value
        depth = work_sheet[f"l{row}"].value
        short_descrition = work_sheet[f"m{row}"].value
        if size != None or width != None or height != None or depth != None:
                
                width = "Width"
                height = "Height"
                depth = "Depth"
                rest.append(width)
                rest.append(height)
                rest.append(depth)
        
        if weight != None:
                rest.append("Weigth")

        if short_descrition != None:
                rest.append("Short_description")

            

        other_columns = rest
        for column in (8, work_sheet.max_column + 1):
            other_columns.append(work_sheet.cell(1,column).value)


        
        
        client_sheet.append(["Manufacturer","Price","Discount","Reference","Name","EAN13","Category"]+rest)
        
        




       


        
        # Client Catalogue




        temp_row = [
            supplier,
            price,
            discount,
            reference,
            name,
            ean,
            category,
        ]



        if width != None or height != None or depth != None:
            temp_row.append(width)
            temp_row.append(height)
            temp_row.append(depth)
        
        if weight != None:
            temp_row.append(weight)
        if short_descrition != None:
            temp_row.append(short_descrition)

        client_sheet.append(temp_row+other_columns)


        # ADS Catalogue

        temp_brand = choose_car_brand(car_brands, name)
        if brand != temp_brand:
            brand = temp_brand
            asd_sheet.append([brand])

            other_columns = []
            for column in (8, work_sheet.max_column + 1):
                other_columns.append(work_sheet.cell(1,column).value)
                # print(1, column)

            asd_sheet.append(["Part Number","Description","Price Exc Vat €"]+other_columns)

        temp_row = [
            reference,
            name,
            price,
        ]


        if width != None or height != None or depth != None:
            temp_row.append(width)
            temp_row.append(height)
            temp_row.append(depth)
        
        if weight != None:
            temp_row.append(weight)

        other_columns = []

        # for column in (14,  20):
    print(work_sheet.max_column, 304)
        #     other_columns.append(price_sheet.cell(row,column).value)










        

  
    
wb.save(manufacturer.upper() + "_update_sheet.xlsx".upper())






