

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s




sheet_name = "Prices - 31 August 2023"
wb = openpyxl.load_workbook("akrapovic_catalogue.xlsx") 
print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


brand_name = None
sheet_2.append(["Manufacturer","Supplier","Price","Purchase","%","VAT","Reference","Product","EAN13","Category","Meta title","Tags","Keywords","Rewrite"])

for row in range(2, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    

    temp_row = [
        "Akrapovic",                                # "manufacturer"
        "Akrapovic",                                # "Supplier"
        sheet_1[f"m{row}"].value,                   # "price"
        Decimal(sheet_1[f"m{row}"].value) * Decimal(0.75), # "purchase"
        "15",                                       # " % "
        "7",                                        # "VAT"
        sheet_1[f"a{row}"].value,                   # "reference"
        sheet_1[f"i{row}"].value + " " + sheet_1[f"j{row}"].value  + " Akrapovic " + sheet_1[f"c{row}"].value, # "product"
        str(sheet_1[f"b{row}"].value),                   # "EAN13"
        "Home",                                     # "category"
        sheet_1[f"i{row}"].value + " " + sheet_1[f"j{row}"].value  + " Akrapovic " + sheet_1[f"c{row}"].value,                   # "Meta title"
        "Akrapovic," + sheet_1[f"a{row}"].value.replace(" ", ",").replace("-", ",").replace("/", ",")  + ","+ sheet_1[f"i{row}"].value.replace(" ", ",").replace("-", ",").replace("/", ",")  + "," + sheet_1[f"j{row}"].value.replace(" ", ",").replace("-", ",").replace("/", ","),                   # "Tags"
        "Akrapovic," + sheet_1[f"a{row}"].value.replace(" ", ",").replace("-", ",").replace("/", ",")  + ","+ sheet_1[f"i{row}"].value.replace(" ", ",").replace("-", ",").replace("/", ",")  + "," + sheet_1[f"j{row}"].value.replace(" ", ",").replace("-", ",").replace("/", ","),                   # "Keywords"
        slugify("Akrapovic-" + sheet_1[f"a{row}"].value),                   # "rewrite"
       
    ]
    sheet_2.append(temp_row)
    
wb.save("akrapovic_all_stars_import_catalogue_new.xlsx")






