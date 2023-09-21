
# generate bue one

import openpyxl
from datetime import date

sheet_name = "Table 1"
wb = openpyxl.load_workbook("racingline_base.xlsx") 
print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


brand_name = None
for row in range(1, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()

    
    if brand_name != sheet_1[f"d{row}"].value and sheet_1[f"d{row}"].value:
        brand_name = str(sheet_1[f"d{row}"].value).upper()
        sheet_2.append([sheet_1[f"d{row}"].value, "", ""])
        sheet_2.append(["Part Number",  "Description", "Retail Price Ex VAT"])

    temp_row = [
        sheet_1[f"a{row}"].value, 
        sheet_1[f"b{row}"].value + " ", 
        sheet_1[f"c{row}"].value, 
    ]
    sheet_2.append(temp_row)
    


wb.save("racingline_catalogue_new.xlsx")


