

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s




sheet_name = "June 2023"
wb = openpyxl.load_workbook("racingline_base.xlsx") 
print(wb.sheetnames) 

sheet_1 = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

sheet_2 = wb.create_sheet(title=f"{sheet_name} - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet


brand_name = None
sheet_2.append(["Manufacturer","Supplier","Price","Purchase","%","VAT","Reference","Name","Category","Meta title","Tags","Keywords","Rewrite"])

for row in range(2, sheet_1.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    if(row<=410):
      temp_row = list()
      print(Decimal(sheet_1[f"c{row}"].value)-Decimal(0.1))
      print(row)

      temp_row = [
          "Racingline",                                # "manufacturer"
          "Racingline",                                # "Supplier"
          sheet_1[f"f{row}"].value,                   # "price"
          sheet_1[f"c{row}"].value,          # "purchase"
          (Decimal(sheet_1[f"d{row}"].value)-Decimal(0.1)) * 100,             # " % "
          "7",                                        # "VAT"
          sheet_1[f"a{row}"].value,                   # "reference"
          sheet_1[f"b{row}"].value + " Racingline " , # "Name"
          "Home",                   # "category"
          "Racingline " + sheet_1[f"b{row}"].value,                   # "Meta title"
          "Racingline," + sheet_1[f"b{row}"].value+"," +sheet_1[f"a{row}"].value,  # "Tags"
          "Racingline," + sheet_1[f"b{row}"].value+"," +sheet_1[f"a{row}"].value,  # "Keywords"
          slugify("Racingline-" + sheet_1[f"a{row}"].value),          # "rewrite"
      ]
      sheet_2.append(temp_row)
    
wb.save("racingline_all_stars_import_catalogue_new.xlsx")






