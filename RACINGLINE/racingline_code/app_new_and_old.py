

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s








def read_excel_file():


    wb = openpyxl.load_workbook("WorkFile.xlsx") 

    out_date_sheet = wb.create_sheet(title=f"out_date - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet

    sheet_name = "website"
    website_sheet = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
    website_references = list()

    for row in range(1, website_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
        if website_sheet[f"c{row}"].value != None: 
            website_references.append(website_sheet[f"c{row}"].value.strip())


    
    new_sheet = wb.create_sheet(title=f"new - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
    
    sheet_name = "prices"
    price_sheet = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
    price_references = list()
    for row in range(1, price_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
        price_references.append(price_sheet[f"g{row}"].value.strip())



    for reference in website_references:
        reference = reference.strip()
        if not reference in price_references:
            out_date_sheet.append([reference])


    for row in range(1, price_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
        reference = price_sheet[f"g{row}"].value.strip()
        if not reference in website_references:
            temp_row = list()
            for col in range(1, price_sheet.max_column + 1):
                # print(row, col)
                cell_value = price_sheet.cell(row, col).value
                temp_row.append(cell_value)
                
            new_sheet.append(temp_row)
    
    wb.save("RACINLINE_UPDATE_SHEET.xlsx")
    
    


read_excel_file()


