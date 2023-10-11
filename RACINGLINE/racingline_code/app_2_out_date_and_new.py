

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s


wb = openpyxl.load_workbook("racingline_base.xlsx") 
# print(wb.sheetnames) 
sheet_name = "website"
website_sheet = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
website_references = list()
out_date_sheet = wb.create_sheet(title=f"out_date - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet

for row in range(1, website_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    if website_sheet[f"c{row}"].value != None: 
      website_references.append(website_sheet[f"c{row}"].value.strip())


sheet_name = "prices"
price_sheet = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.
price_references = list()
new_sheet = wb.create_sheet(title=f"new - {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet

for row in range(2, 410): # To iterate over all the row and column of the sheet and get each value.
    price_references.append(price_sheet[f"a{row}"].value.strip())



for reference in website_references:
    reference = reference.strip()
    if not reference in price_references:
        out_date_sheet.append([reference])


for reference in price_references:
    reference = reference.strip()
    if not reference in website_references:
        new_sheet.append([reference])

  
    
wb.save("racinline_update_sheet.xlsx")






