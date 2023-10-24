

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s




wb = openpyxl.load_workbook("Work_FIle.xlsx") 



sheet_name_variations = "Folha2"

# print(wb.sheetnames) 

sheet_variations = wb[sheet_name_variations] # To accsses the a sheet in the workbook. And create a sheet object.

variariations = dict()
reference = ""
for row in range(2, sheet_variations.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    if reference != sheet_variations[f"a{row}"].value:
        reference = sheet_variations[f"a{row}"].value.strip()
        variariations[reference] = list()
        variariations[reference].append(sheet_variations[f"b{row}"].value)
    else:
        variariations[reference].append(sheet_variations[f"b{row}"].value)
    



# print(variariations['AC-500-AH-Z'])

# for x in variariations:
#   print(type(variariations[x]))


sheet_name = "Folha1"


working_sheet = wb[sheet_name] # To accsses the a sheet in the workbook. And create a sheet object.

new_sheet = wb.create_sheet(title=f"new-sheet- {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet
on_date = list()
for row in range(2, working_sheet.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    reference = working_sheet[f"a{row}"].value.strip()
    # print(working_sheet[f"a{row}"].value)
    
    if variariations.get(reference) != None:
        
        for variation in variariations.get(reference):
            print(variation)
            # print(variation)
            new_sheet.append([
                working_sheet[f"a{row}"].value,
                working_sheet[f"b{row}"].value,
                working_sheet[f"c{row}"].value,
                working_sheet[f"d{row}"].value,
                working_sheet[f"e{row}"].value,
                variation,
            ])
    else:
        # print(working_sheet[f"a{row}"].value)
        new_sheet.append([
            working_sheet[f"a{row}"].value,
            working_sheet[f"b{row}"].value,
            working_sheet[f"c{row}"].value,
            working_sheet[f"d{row}"].value,
            working_sheet[f"e{row}"].value,
        ])   
    
wb.save("alfa_with_variations.xlsx")






