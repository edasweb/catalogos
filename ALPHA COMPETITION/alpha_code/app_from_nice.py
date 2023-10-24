

# generate bue one

import openpyxl
from datetime import date
import re
from decimal import Decimal
def slugify(s):
#   s = s.lower().strip()
  s = re.sub(r'[^\w\s-]', '', s)
  s = re.sub(r'[\s_-]+', '-', s)
  s = re.sub(r'^-+|-+$', '', s)
  return s




wb = openpyxl.load_workbook("alfa_with_variations3.xlsx") 
sheet_price_name = "prices"
nice_sheet = wb.create_sheet(title=f"nice- {date.today().strftime('%b-%d-%Y')}", index=0) # Use the ".create_sheet()" method to create a new sheet in the workbook. The index parameter is the position of the sheet

# print(wb.sheetnames) 

sheet_prices = wb[sheet_price_name] # To accsses the a sheet in the workbook. And create a sheet object.

prices = {}
references = list()
for row in range(2, sheet_prices.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    reference = sheet_prices[f"h{row}"].value.upper().strip().replace(" ", "")
    

    if prices.get(reference) == None and reference != None:
        references.append(reference)
        prices[reference] = {
            "price":sheet_prices[f"c{row}"].value, 
                             "wholesale_price":sheet_prices[f"d{row}"].value,
                             "attr_price":sheet_prices[f"e{row}"].value,
                             "attr_wholesale_price":sheet_prices[f"f{row}"].value,
                             "name":sheet_prices[f"g{row}"].value,
                             }



sheet_name_brand_together = "brand-together"




sheet_brand_together = wb[sheet_name_brand_together] # To accsses the a sheet in the workbook. And create a sheet object.

brand_name = None

for row in range(1, sheet_brand_together.max_row + 1): # To iterate over all the row and column of the sheet and get each value.
    temp_row = list()
    
    if brand_name !=str(sheet_brand_together[f"b{row}"].value).upper().strip():
        brand_name = str(sheet_brand_together[f"b{row}"].value).upper().strip()
        nice_sheet.append([brand_name])
        nice_sheet.append(["Part Number", "Description", "Retail Price Ex VAT €"
])
    
    reference = sheet_brand_together[f"f{row}"].value.upper().strip().replace(" ", "")
    if not reference in references:
        pass
        # print("row: ", row, " ", reference)
      
    if reference in references:
        temp_row = [
            reference,
            f'{sheet_brand_together[f"b{row}"].value} {sheet_brand_together[f"c{row}"].value} {prices.get(reference)["name"]}', 
            f'{prices[reference]["price"]}', 
        ]
    
    nice_sheet.append(temp_row)
    

    
wb.save("alfa_nice_2.xlsx")






